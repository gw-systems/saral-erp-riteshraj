import logging
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from integrations.tallysync.services.financial_analytics_service import FinancialAnalyticsService

logger = logging.getLogger(__name__)
from integrations.tallysync.services.project_analytics_service import ProjectAnalyticsService
from integrations.tallysync.services.cash_flow_service import CashFlowService
from integrations.tallysync.services.gst_service import GSTService
from integrations.tallysync.services.salesperson_analytics_service import SalespersonAnalyticsService
from integrations.tallysync.services.client_analytics_service import ClientAnalyticsService
from integrations.tallysync.services.vendor_analytics_service import VendorAnalyticsService
from integrations.tallysync.services.ledger_analytics_service import LedgerAnalyticsService
from integrations.tallysync.services.aging_report_service import AgingReportService
from integrations.tallysync.models import TallyCompany
from datetime import datetime, date
from decimal import Decimal
import json
from django.utils import timezone
from integrations.tallysync.snapshot_models import ProjectFinancialSnapshot
from projects.models import ProjectCode


class DecimalEncoder(json.JSONEncoder):
    """Custom JSON encoder for Decimal objects"""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, datetime):
            return obj.isoformat()
        if isinstance(obj, date):
            return obj.isoformat()
        return super(DecimalEncoder, self).default(obj)
    
def format_month_year(date_obj):
    """Format date as 'Jan 2024' for frontend"""
    if isinstance(date_obj, str):
        date_obj = datetime.strptime(date_obj, '%Y-%m-%d').date()
    return date_obj.strftime('%b %Y')


@login_required
@require_http_methods(["GET"])
def api_executive_summary(request):
    """API: Executive financial summary"""
    
    # Get filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    # Parse dates
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get data
    service = FinancialAnalyticsService(start_date, end_date, company_id)
    data = service.get_executive_summary()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_revenue_breakdown(request):
    """API: Revenue breakdown by type"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = FinancialAnalyticsService(start_date, end_date, company_id)
    data = service.get_revenue_breakdown()
    
    return JsonResponse({'breakdown': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_company_summary(request):
    """API: Company-wise financial summary"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = FinancialAnalyticsService(start_date, end_date)
    data = service.get_company_wise_summary()
    
    return JsonResponse({'companies': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_monthly_trend(request):
    """API: Monthly financial trend"""
    
    months = int(request.GET.get('months', 6))
    company_id = request.GET.get('company_id')
    
    service = FinancialAnalyticsService(company_id=company_id)
    data = service.get_monthly_trend(months=months)
    
    return JsonResponse({'trend': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_project_profitability(request):
    """API: All projects profitability summary.
    Supports optional client_names/vendor_names filters for drill-down.
    """

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    client_names = request.GET.get('client_names')  # comma-separated
    vendor_names = request.GET.get('vendor_names')  # comma-separated

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    client_list = [n.strip() for n in client_names.split(',')] if client_names else None
    vendor_list = [n.strip() for n in vendor_names.split(',')] if vendor_names else None

    service = ProjectAnalyticsService(start_date, end_date)
    data = service.get_all_projects_summary(client_names=client_list, vendor_names=vendor_list)

    return JsonResponse({'projects': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_client_profitability(request):
    """API: Client-level profitability summary"""

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    service = ClientAnalyticsService(start_date, end_date)
    data = service.get_all_clients_summary()

    return JsonResponse({'clients': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_vendor_profitability(request):
    """API: Vendor-level profitability summary"""

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    service = VendorAnalyticsService(start_date, end_date)
    data = service.get_all_vendors_summary()

    return JsonResponse({'vendors': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_project_detail(request, project_id):
    """API: Single project detail"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = ProjectAnalyticsService(start_date, end_date)
    data = service.get_project_detail(project_id)
    
    return JsonResponse(data, encoder=DecimalEncoder)


def _parse_inventory_items(raw_xml):
    """Extract service/stock items with descriptions from raw Tally XML."""
    import re
    from html import unescape
    items = []
    if not raw_xml:
        return items
    blocks = re.findall(
        r'<(?:ALL)?INVENTORYENTRIES\.LIST>(.*?)</(?:ALL)?INVENTORYENTRIES\.LIST>',
        raw_xml, re.DOTALL
    )
    for block in blocks:
        name_m = re.search(r'<STOCKITEMNAME[^>]*>([^<]+)</', block)
        if not name_m:
            continue
        name = unescape(name_m.group(1).strip())
        descriptions = re.findall(r'<BASICUSERDESCRIPTION>([^<]+)</', block)
        amount_m = re.search(r'<AMOUNT[^>]*>([^<]+)</', block)
        # Clean up descriptions: unescape, strip whitespace, remove blank lines
        clean_descs = []
        for d in descriptions:
            d = unescape(d).strip().replace('\r\n', ' ').replace('\n', ' ')
            d = ' '.join(d.split())  # collapse internal whitespace
            if d:
                clean_descs.append(d)
        items.append({
            'service': name,
            'descriptions': clean_descs,
            'amount': amount_m.group(1).strip() if amount_m else '',
        })
    return items


@login_required
@require_http_methods(["GET"])
def api_voucher_detail(request, voucher_id):
    """API: Voucher detail with ledger entries, bill references, and service items."""
    from integrations.tallysync.models import TallyVoucher

    try:
        voucher = TallyVoucher.objects.select_related('company').get(pk=voucher_id)
    except TallyVoucher.DoesNotExist:
        return JsonResponse({'error': 'Voucher not found'}, status=404)

    ledger_entries = list(
        voucher.ledger_entries.all().values(
            'ledger_name', 'amount', 'is_debit',
        )
    )

    from integrations.tallysync.models import TallyBillReference
    bill_refs = list(
        TallyBillReference.objects.filter(
            ledger_entry__voucher=voucher
        ).values('bill_name', 'bill_type', 'amount')
    )
    bill_refs = [
        {'bill_number': b['bill_name'], 'bill_type': b['bill_type'], 'amount': b['amount']}
        for b in bill_refs
    ]

    inventory_items = _parse_inventory_items(voucher.raw_xml)

    return JsonResponse({
        'id': voucher.id,
        'voucher_type': voucher.voucher_type,
        'voucher_number': voucher.voucher_number,
        'date': voucher.date.isoformat(),
        'amount': voucher.amount,
        'party_name': voucher.party_ledger_name,
        'party_display_name': voucher.party_name or voucher.party_ledger_name,
        'party_gstin': voucher.party_gstin,
        'party_state': voucher.party_state,
        'narration': voucher.narration,
        'reference': voucher.reference,
        'company': voucher.company.name if voucher.company else '',
        'buyer_name': voucher.buyer_name,
        'buyer_gstin': voucher.buyer_gstin,
        'buyer_state': voucher.buyer_state,
        'consignee_name': voucher.consignee_name,
        'consignee_gstin': voucher.consignee_gstin,
        'payment_mode': voucher.payment_mode,
        'cheque_number': voucher.cheque_number,
        'cheque_date': voucher.cheque_date.isoformat() if voucher.cheque_date else '',
        'eway_bill_number': voucher.eway_bill_number,
        'billing_month': voucher.billing_month,
        'need_to_pay': voucher.need_to_pay,
        'remark': voucher.remark,
        'credit_period': voucher.credit_period,
        'transaction_type': voucher.transaction_type,
        'utr_number': voucher.utr_number,
        'is_cancelled': voucher.is_cancelled,
        'ledger_entries': ledger_entries,
        'bill_references': bill_refs,
        'inventory_items': inventory_items,
    }, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_project_lifecycle(request, project_id):
    """API: Project lifecycle analysis with monthly breakdown"""

    service = ProjectAnalyticsService()
    data = service.get_project_lifecycle_analysis(project_id)

    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_top_projects(request):
    """API: Top profitable projects"""
    
    limit = int(request.GET.get('limit', 10))
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = ProjectAnalyticsService(start_date, end_date)
    data = service.get_top_profitable_projects(limit=limit)
    
    return JsonResponse({'projects': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_cash_flow_summary(request):
    """API: Cash flow summary"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = CashFlowService(start_date, end_date, company_id)
    data = service.get_cash_summary()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_cash_flow_trend(request):
    """API: Cash flow trend"""
    
    months = int(request.GET.get('months', 6))
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = CashFlowService(start_date, end_date, company_id)
    data = service.get_cash_flow_trend(months=months)
    
    return JsonResponse({'trend': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_receivables(request):
    """API: Receivables summary"""
    
    company_id = request.GET.get('company_id')
    
    service = CashFlowService(company_id=company_id)
    data = service.get_receivables_summary()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_payables(request):
    """API: Payables summary"""
    
    company_id = request.GET.get('company_id')
    
    service = CashFlowService(company_id=company_id)
    data = service.get_payables_summary()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_receivables_aging(request):
    """API: Party-wise receivables aging report"""
    company_id = request.GET.get('company_id')
    active_only = request.GET.get('active_only', 'true').lower() != 'false'
    service = AgingReportService(company_id=company_id)
    data = service.get_receivables_aging(active_only=active_only)
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_payables_aging(request):
    """API: Party-wise payables aging report"""
    company_id = request.GET.get('company_id')
    service = AgingReportService(company_id=company_id)
    data = service.get_payables_aging()
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_party_aging_detail(request):
    """API: Voucher-level detail for a specific party's aging"""
    party_name = request.GET.get('party_name', '')
    report_type = request.GET.get('report_type', 'receivables')
    company_id = request.GET.get('company_id')

    if not party_name:
        return JsonResponse({'error': 'party_name is required'}, status=400)

    service = AgingReportService(company_id=company_id)
    data = service.get_party_detail(party_name, report_type)
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_project_aging_download(request):
    """Download project-wise aging as Excel."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from django.http import HttpResponse

    report_type = request.GET.get('report_type', 'receivables')
    company_id = request.GET.get('company_id')
    service = AgingReportService(company_id=company_id)
    rows = service.get_project_wise_aging(report_type=report_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.capitalize()} Aging - Project"

    # Header
    headers = ['Project', 'Party', 'Bill Name', 'Invoice Date', 'Days Old', 'Bucket', 'Invoiced', 'Settled', 'Outstanding']
    header_fill = PatternFill('solid', fgColor='1E40AF' if report_type == 'receivables' else 'BE123C')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    bucket_colors = {'0-30': 'D1FAE5', '31-60': 'FEF9C3', '61-90': 'FFEDD5', '90+': 'FEE2E2'}

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row['project_name'])
        ws.cell(row=i, column=2, value=row['party_name'])
        ws.cell(row=i, column=3, value=row['bill_name'])
        date_cell = ws.cell(row=i, column=4, value=row['date'])
        if row['date']:
            date_cell.number_format = 'DD-MM-YYYY'
        ws.cell(row=i, column=5, value=row['days_old'])
        bucket_cell = ws.cell(row=i, column=6, value=row['bucket'])
        bucket_cell.fill = PatternFill('solid', fgColor=bucket_colors.get(row['bucket'], 'FFFFFF'))
        for col, key in [(7, 'invoiced'), (8, 'settled'), (9, 'outstanding')]:
            c = ws.cell(row=i, column=col, value=float(row[key]))
            c.number_format = '#,##0.00'
        if row['outstanding'] > 0 and row['bucket'] == '90+':
            ws.cell(row=i, column=9).font = Font(bold=True, color='DC2626')

    # Totals row
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for col, key in [(7, 'invoiced'), (8, 'settled'), (9, 'outstanding')]:
        total = sum(float(r[key]) for r in rows)
        c = ws.cell(row=total_row, column=col, value=total)
        c.number_format = '#,##0.00'
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='E0E7FF' if report_type == 'receivables' else 'FFE4E6')

    # Column widths
    for col, width in zip(range(1, 10), [30, 35, 20, 14, 10, 10, 15, 15, 15]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from django.utils import timezone
    filename = f"aging_{report_type}_project_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_http_methods(["GET"])
def api_party_aging_download(request):
    """Download party-wise aging as Excel."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    report_type = request.GET.get('report_type', 'receivables')
    company_id = request.GET.get('company_id')
    service = AgingReportService(company_id=company_id)
    # active_only only applies to receivables (filter by active ERP clients)
    # payables are vendors — never filtered by active project
    if report_type == 'receivables':
        active_only = request.GET.get('active_only', 'true').lower() != 'false'
        result = service.get_receivables_aging(active_only=active_only)
    else:
        result = service.get_payables_aging()
    parties = result['parties']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.capitalize()} Aging - Party"

    headers = ['Party Name', 'Outstanding', '0-30 Days', '31-60 Days', '61-90 Days', '90+ Days', 'Oldest (Days)', 'Open Bills']
    header_fill = PatternFill('solid', fgColor='1E40AF' if report_type == 'receivables' else 'BE123C')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, p in enumerate(parties, 2):
        ws.cell(row=i, column=1, value=p['party_name'])
        for col, key in [(2, 'outstanding'), (3, 'aging_0_30'), (4, 'aging_31_60'), (5, 'aging_61_90'), (6, 'aging_90_plus')]:
            c = ws.cell(row=i, column=col, value=float(p[key]))
            c.number_format = '#,##0.00'
            if key == 'aging_90_plus' and float(p[key]) > 0:
                c.font = Font(bold=True, color='DC2626')
        ws.cell(row=i, column=7, value=p['days_oldest'])
        ws.cell(row=i, column=8, value=p['bill_count'])

    # Totals row
    total_row = len(parties) + 2
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for col, key in [(2, 'outstanding'), (3, 'aging_0_30'), (4, 'aging_31_60'), (5, 'aging_61_90'), (6, 'aging_90_plus')]:
        total = sum(float(p[key]) for p in parties)
        c = ws.cell(row=total_row, column=col, value=total)
        c.number_format = '#,##0.00'
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='E0E7FF' if report_type == 'receivables' else 'FFE4E6')

    for col, width in zip(range(1, 9), [40, 15, 12, 12, 12, 12, 14, 10]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from django.utils import timezone
    filename = f"aging_{report_type}_party_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_http_methods(["GET"])
def api_aging_summary(request):
    """API: Combined aging summary (receivables + payables)"""
    company_id = request.GET.get('company_id')
    service = AgingReportService(company_id=company_id)
    data = service.get_aging_summary()
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_companies(request):
    """API: List all Tally companies for filter dropdowns"""
    companies = list(TallyCompany.objects.values('id', 'name').order_by('name'))
    return JsonResponse({'companies': companies})


@login_required
@require_http_methods(["POST"])
def api_discover_companies(request):
    """API: Discover/refresh companies from Tally server"""
    from integrations.tallysync.services.sync_service import TallySyncService
    try:
        service = TallySyncService()
        result = service.sync_companies(triggered_by_user=request.user.username)
        companies = list(TallyCompany.objects.values('id', 'name').order_by('name'))
        return JsonResponse({'success': True, 'companies': companies, 'detail': result})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def api_gst_summary(request):
    """API: GST summary"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = GSTService(start_date, end_date, company_id)
    data = service.get_gst_summary()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_gst_monthly_return(request):
    """API: GST monthly return data"""
    
    month = int(request.GET.get('month'))
    year = int(request.GET.get('year'))
    company_id = request.GET.get('company_id')
    
    service = GSTService(company_id=company_id)
    data = service.get_monthly_gst_return_data(month, year)
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_gst_by_state(request):
    """API: GST breakdown by state"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = GSTService(start_date, end_date)
    data = service.get_gst_by_state()
    
    return JsonResponse({'states': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_salesperson_performance(request):
    """API: Salesperson performance summary"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = SalespersonAnalyticsService(start_date, end_date)
    data = service.get_all_salesperson_summary()
    
    return JsonResponse({'salespeople': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_salesperson_detail(request, salesperson_name):
    """API: Single salesperson detail"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = SalespersonAnalyticsService(start_date, end_date)
    data = service.get_salesperson_detail(salesperson_name)
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_tds_summary(request):
    """API: TDS summary"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = LedgerAnalyticsService(start_date, end_date, company_id)
    data = service.get_tds_summary()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_bank_transactions(request):
    """API: Bank transactions summary"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = LedgerAnalyticsService(start_date, end_date, company_id)
    data = service.get_bank_transactions_summary()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_vendor_expenses(request):
    """API: Vendor-wise expenses"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = LedgerAnalyticsService(start_date, end_date, company_id)
    data = service.get_vendor_expense_summary()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_vendor_detail(request, vendor_name):
    """API: Single vendor detail — project breakdown + all vouchers"""

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    service = LedgerAnalyticsService(start_date, end_date, company_id)
    data = service.get_vendor_detail(vendor_name)

    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_customer_revenue(request):
    """API: Customer-wise revenue"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = LedgerAnalyticsService(start_date, end_date, company_id)
    data = service.get_customer_revenue_summary()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_detailed_gst_breakdown(request):
    """API: Detailed GST breakdown using actual field values"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = GSTService(start_date, end_date, company_id)
    data = service.get_detailed_gst_breakdown()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_payment_mode_analysis(request):
    """API: Enhanced payment mode analysis with cheque details"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = LedgerAnalyticsService(start_date, end_date, company_id)
    data = service.get_payment_mode_analysis()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_ledger_groups_summary(request):
    """API: All ledger groups summary"""
    
    from integrations.tallysync.services.ledger_group_service import LedgerGroupService
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = LedgerGroupService(start_date, end_date, company_id)
    data = service.get_all_groups_summary()
    
    return JsonResponse({'groups': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_income_statement_groups(request):
    """API: Income statement organized by groups"""
    
    from integrations.tallysync.services.ledger_group_service import LedgerGroupService
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = LedgerGroupService(start_date, end_date, company_id)
    data = service.get_income_statement_groups()
    
    return JsonResponse(data, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_group_hierarchy(request):
    """API: Ledger group hierarchy with financials"""
    
    from integrations.tallysync.services.ledger_group_service import LedgerGroupService
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    service = LedgerGroupService(start_date, end_date, company_id)
    data = service.get_group_hierarchy()
    
    return JsonResponse({'hierarchy': data}, encoder=DecimalEncoder)


@login_required
@require_http_methods(["GET"])
def api_salesperson_financial_summary(request):
    """
    Get Tally financial summary for logged-in sales manager
    
    Query Parameters:
    - start_date: YYYY-MM-DD (default: previous month start)
    - end_date: YYYY-MM-DD (default: previous month end)
    
    Returns: 
    - Summary financial data
    - Top 5 profitable projects
    - Top 5 loss-making projects  
    - High outstanding projects (>50K)
    - All projects list
    """
    try:
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        # Get salesperson name (from query param for testing, or from user)
        salesperson_name = request.GET.get('sales_manager') or request.user.get_full_name()
        
        if not salesperson_name:
            return JsonResponse({
                'error': 'User profile not complete'
            }, status=400)
        
        # Parse date parameters (default to previous month)
        today = timezone.now().date()
        first_day_current_month = today.replace(day=1)
        last_day_previous_month = first_day_current_month - timedelta(days=1)
        first_day_previous_month = last_day_previous_month.replace(day=1)
        
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': 'Invalid start_date format. Use YYYY-MM-DD'}, status=400)
        else:
            start_date = first_day_previous_month
        
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': 'Invalid end_date format. Use YYYY-MM-DD'}, status=400)
        else:
            end_date = last_day_previous_month
        
        # Validate date range
        if start_date > end_date:
            return JsonResponse({'error': 'start_date cannot be after end_date'}, status=400)
        
        # Get their projects
        projects = ProjectCode.objects.filter(
            sales_manager=salesperson_name,
            project_status__in=['Active', 'Operation Not Started', 'Notice Period']
        )
        
        if not projects.exists():
            return JsonResponse({
                'summary': {
                    'total_revenue': 0,
                    'total_expenses': 0,
                    'total_profit': 0,
                    'avg_margin_pct': 0,
                    'total_outstanding': 0,
                    'project_count': 0,
                },
                'top_profitable': [],
                'top_loss_making': [],
                'high_outstanding': [],
                'projects': [],
                'date_range': {
                    'start_date': start_date.isoformat(),
                    'end_date': end_date.isoformat()
                }
            })
        
        # Get latest snapshot per project in bulk (2 queries max instead of 2N)
        project_ids = list(projects.values_list('project_id', flat=True))

        # First try: snapshots within date range
        in_range = ProjectFinancialSnapshot.objects.filter(
            project_id__in=project_ids,
            snapshot_date__gte=start_date,
            snapshot_date__lte=end_date,
        ).select_related('project').order_by('project_id', '-snapshot_date').distinct('project_id')

        snapshot_map = {s.project_id: s for s in in_range}

        # For projects without in-range snapshot, get latest available
        missing_ids = [pid for pid in project_ids if pid not in snapshot_map]
        if missing_ids:
            fallback = ProjectFinancialSnapshot.objects.filter(
                project_id__in=missing_ids,
            ).select_related('project').order_by('project_id', '-snapshot_date').distinct('project_id')
            for s in fallback:
                snapshot_map[s.project_id] = s

        latest_snapshots = list(snapshot_map.values())
        
        # Build project list
        project_data = []
        total_revenue = Decimal('0')
        total_expenses = Decimal('0')
        total_profit = Decimal('0')
        total_outstanding = Decimal('0')
        
        for snapshot in latest_snapshots:
            project_data.append({
                'project_code': snapshot.project.project_code,  # Full project code
                'tally_sales': snapshot.tally_revenue,  # Sales vouchers
                'tally_purchase': snapshot.tally_expenses,  # Purchase vouchers (rename for clarity)
                'tally_profit': snapshot.tally_profit,
                'tally_margin_pct': snapshot.tally_margin_pct,
                'outstanding': snapshot.outstanding,
                'revenue_per_sqft': snapshot.revenue_per_sqft,
                'profit_per_sqft': snapshot.profit_per_sqft,
                'sales_count': snapshot.sales_count,
                'last_updated': snapshot.last_updated.isoformat(),
            })
            
            total_revenue += snapshot.tally_revenue
            total_expenses += snapshot.tally_expenses
            total_profit += snapshot.tally_profit
            total_outstanding += snapshot.outstanding
        
        avg_margin_pct = (total_profit / total_revenue * 100) if total_revenue > 0 else Decimal('0')
        
        # Sort projects by profit for top/bottom performers
        sorted_by_profit = sorted(project_data, key=lambda x: x['tally_profit'], reverse=True)
        
        # Top 5 profitable projects - with full project_code
        top_profitable = []
        for snapshot in sorted(latest_snapshots, key=lambda x: x.tally_profit, reverse=True)[:5]:
            if snapshot.tally_profit > 0:
                top_profitable.append({
                    'project_code': snapshot.project.project_code,  # Full format from DB
                    'project_id': snapshot.project.project_id,
                    'profit': snapshot.tally_profit,
                    'margin_pct': snapshot.tally_margin_pct
                })

        # Top 5 loss-making projects - with full project_code
        top_loss_making = []
        for snapshot in sorted(latest_snapshots, key=lambda x: x.tally_profit)[:5]:
            if snapshot.tally_profit < 0:
                top_loss_making.append({
                    'project_code': snapshot.project.project_code,  # Full format from DB
                    'project_id': snapshot.project.project_id,
                    'loss': abs(snapshot.tally_profit),
                    'margin_pct': snapshot.tally_margin_pct
                })

        # High outstanding projects (>50K) - with full project_code
        high_outstanding = []
        for snapshot in sorted(latest_snapshots, key=lambda x: x.outstanding, reverse=True):
            if snapshot.outstanding > 50000:
                high_outstanding.append({
                    'project_code': snapshot.project.project_code,  # Full format from DB
                    'project_id': snapshot.project.project_id,
                    'outstanding': snapshot.outstanding
                })
        
        response_data = {
            'summary': {
                'total_revenue': total_revenue,
                'total_expenses': total_expenses,
                'total_profit': total_profit,
                'avg_margin_pct': avg_margin_pct,
                'total_outstanding': total_outstanding,
                'project_count': len(project_data),
            },
            'top_profitable': top_profitable,
            'top_loss_making': top_loss_making,
            'high_outstanding': high_outstanding,
            'projects': project_data,
            'date_range': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat()
            }
        }
        
        return JsonResponse(response_data, encoder=DecimalEncoder)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': str(e)
        }, status=500)
    

@login_required
@require_http_methods(["GET"])
def api_chart_data_monthly_trend(request):
    """
    API: Monthly trend formatted specifically for Chart.js
    Returns data in format: {labels: [], datasets: []}
    """
    
    months = int(request.GET.get('months', 6))
    company_id = request.GET.get('company_id')
    
    service = FinancialAnalyticsService(company_id=company_id)
    trend_data = service.get_monthly_trend(months=months)
    
    # Format for Chart.js
    labels = []
    revenue_data = []
    expense_data = []
    profit_data = []
    
    for item in trend_data:
        # Format month as "Jan 2024"
        month_date = item['month']
        labels.append(month_date.strftime('%b %Y'))
        revenue_data.append(float(item['revenue']))
        expense_data.append(float(item['expenses']))
        profit_data.append(float(item['profit']))
    
    response_data = {
        'labels': labels,
        'datasets': {
            'revenue': revenue_data,
            'expenses': expense_data,
            'profit': profit_data
        }
    }
    
    return JsonResponse(response_data, encoder=DecimalEncoder)

@login_required
@require_http_methods(["GET", "POST"])
def tally_connection_test(request):
    """
    Admin-only Tally connection test page
    GET: Show current configuration
    POST: Run connection test and return results
    """
    # Check if user is admin or super_user
    if request.user.role not in ['admin', 'super_user']:
        return JsonResponse({
            'error': 'Access denied. Admin or Super User role required.'
        }, status=403)

    if request.method == 'POST':
        # Run connection test
        from integrations.tallysync.services.tally_connector_new import TallyConnector

        connector = TallyConnector()
        result = connector.test_connection()

        # Add configuration details
        result['configuration'] = {
            'host': connector.host,
            'port': connector.port,
            'timeout': connector.timeout,
            'base_url': connector.base_url,
        }

        return JsonResponse(result, encoder=DecimalEncoder)

    else:
        # GET: Return configuration info for display
        connector = TallyConnector()
        return JsonResponse({
            'configuration': {
                'host': connector.host,
                'port': connector.port,
                'timeout': connector.timeout,
                'company_name': getattr(settings, 'TALLY_COMPANY_NAME', ''),
            }
        }, encoder=DecimalEncoder)


@login_required
@require_http_methods(["POST"])
def api_network_diagnostic(request):
    """Comprehensive network diagnostic for Tally connectivity from this server."""
    if request.user.role not in ['admin', 'super_user']:
        return JsonResponse({'error': 'Access denied.'}, status=403)

    import socket
    import time
    import subprocess
    import urllib.request
    import requests as req_lib

    # Use same source as TallyConnector
    from integrations.tallysync.models import TallySyncSettings
    db_settings = TallySyncSettings.load()
    tunnel_url = (db_settings.tunnel_url or '').strip()
    host = db_settings.server_ip or getattr(settings, 'TALLY_HOST', 'localhost')
    port = int(db_settings.server_port or getattr(settings, 'TALLY_PORT', 2245))

    tests = []

    def add(name, status, detail='', ms=None):
        entry = {'test': name, 'status': status, 'detail': detail}
        if ms is not None:
            entry['ms'] = round(ms, 1)
        tests.append(entry)

    # ── Test 1: Environment info ──
    import os
    env = os.environ.get('ENVIRONMENT', os.environ.get('ENV', 'unknown'))
    add('environment', 'info', f'ENV={env}, host={host}, port={port}')

    # ── Test 2: Outbound public IP ──
    try:
        t0 = time.time()
        resp = urllib.request.urlopen('https://api.ipify.org', timeout=5)
        ms = (time.time() - t0) * 1000
        ip = resp.read().decode().strip()
        add('outbound_ip', ip, 'Public IP this server uses for outbound connections', ms)
    except Exception as e:
        add('outbound_ip', 'ERROR', str(e))

    # ── Test 3: DNS resolution ──
    try:
        t0 = time.time()
        addrs = socket.getaddrinfo(host, port, socket.AF_INET, socket.SOCK_STREAM)
        ms = (time.time() - t0) * 1000
        resolved = list(set(a[4][0] for a in addrs))
        add('dns_resolve', 'OK', f'{host} → {resolved}', ms)
    except socket.gaierror as e:
        add('dns_resolve', 'FAIL', f'Cannot resolve {host}: {e}')
    except Exception as e:
        add('dns_resolve', 'FAIL', str(e))

    # ── Test 4: Outbound TCP to google.com:443 (sanity check) ──
    try:
        t0 = time.time()
        sock = socket.create_connection(('google.com', 443), timeout=5)
        ms = (time.time() - t0) * 1000
        sock.close()
        add('tcp_sanity_google_443', 'OK', 'Outbound TCP works', ms)
    except Exception as e:
        add('tcp_sanity_google_443', 'FAIL', f'Outbound TCP broken: {e}')

    # ── Test 5: TCP to Tally host on port 80 (is host reachable at all?) ──
    try:
        t0 = time.time()
        sock = socket.create_connection((host, 80), timeout=5)
        ms = (time.time() - t0) * 1000
        sock.close()
        add(f'tcp_{host}_80', 'OK', 'Host reachable on port 80', ms)
    except socket.timeout:
        add(f'tcp_{host}_80', 'TIMEOUT', 'Host not reachable on port 80 (5s)')
    except ConnectionRefusedError:
        add(f'tcp_{host}_80', 'REFUSED', 'Port 80 closed — but host is reachable (got RST)')
    except OSError as e:
        add(f'tcp_{host}_80', 'ERROR', str(e))

    # ── Test 6: TCP to Tally host on port 443 ──
    try:
        t0 = time.time()
        sock = socket.create_connection((host, 443), timeout=5)
        ms = (time.time() - t0) * 1000
        sock.close()
        add(f'tcp_{host}_443', 'OK', 'Host reachable on port 443', ms)
    except socket.timeout:
        add(f'tcp_{host}_443', 'TIMEOUT', 'Host not reachable on port 443 (5s)')
    except ConnectionRefusedError:
        add(f'tcp_{host}_443', 'REFUSED', 'Port 443 closed — but host is reachable (got RST)')
    except OSError as e:
        add(f'tcp_{host}_443', 'ERROR', str(e))

    # ── Test 7: Raw TCP to Tally port (THE critical test) ──
    for timeout_s in [5, 10]:
        try:
            t0 = time.time()
            sock = socket.create_connection((host, port), timeout=timeout_s)
            ms = (time.time() - t0) * 1000
            sock.close()
            add(f'tcp_{host}_{port}_t{timeout_s}s', 'OK', f'Tally port reachable', ms)
        except socket.timeout:
            add(f'tcp_{host}_{port}_t{timeout_s}s', 'TIMEOUT', f'Port {port} unreachable ({timeout_s}s) — firewall dropping packets')
        except ConnectionRefusedError:
            add(f'tcp_{host}_{port}_t{timeout_s}s', 'REFUSED', f'Port {port} actively refused (RST) — Tally not listening or port wrong')
        except OSError as e:
            add(f'tcp_{host}_{port}_t{timeout_s}s', 'ERROR', str(e))

    # ── Test 8: HTTP GET to Tally (what the actual connector does) ──
    try:
        t0 = time.time()
        r = req_lib.get(f'http://{host}:{port}', timeout=10)
        ms = (time.time() - t0) * 1000
        body_preview = r.text[:200]
        is_tally = 'Tally' in r.text
        add('http_get_tally', 'OK' if is_tally else 'UNEXPECTED',
            f'HTTP {r.status_code}, Tally={"yes" if is_tally else "no"}, body={body_preview}', ms)
    except req_lib.exceptions.ConnectTimeout:
        add('http_get_tally', 'CONNECT_TIMEOUT', f'TCP connect timed out (10s) — cannot reach {host}:{port}')
    except req_lib.exceptions.ReadTimeout:
        add('http_get_tally', 'READ_TIMEOUT', 'TCP connected but no HTTP response (10s) — port open but not HTTP?')
    except req_lib.exceptions.ConnectionError as e:
        err_str = str(e)
        if 'Connection refused' in err_str:
            add('http_get_tally', 'REFUSED', f'Connection refused at {host}:{port}')
        elif 'Name or service not known' in err_str:
            add('http_get_tally', 'DNS_FAIL', f'Cannot resolve {host}')
        else:
            add('http_get_tally', 'CONN_ERROR', err_str[:300])
    except Exception as e:
        add('http_get_tally', 'ERROR', f'{type(e).__name__}: {e}')

    # ── Test 9: Traceroute (first 10 hops, non-blocking, 15s max) ──
    try:
        result = subprocess.run(
            ['traceroute', '-n', '-m', '10', '-w', '2', host],
            capture_output=True, text=True, timeout=15
        )
        lines = result.stdout.strip().split('\n') if result.stdout else []
        add('traceroute', 'INFO', '\n'.join(lines[:12]))
    except FileNotFoundError:
        # traceroute not installed — try tracepath
        try:
            result = subprocess.run(
                ['tracepath', '-n', '-m', '10', host],
                capture_output=True, text=True, timeout=15
            )
            lines = result.stdout.strip().split('\n') if result.stdout else []
            add('tracepath', 'INFO', '\n'.join(lines[:12]))
        except Exception:
            add('traceroute', 'SKIP', 'traceroute/tracepath not available in container')
    except subprocess.TimeoutExpired:
        add('traceroute', 'TIMEOUT', 'Traceroute timed out after 15s')
    except Exception as e:
        add('traceroute', 'ERROR', str(e))

    # ── Test 10: Ping (ICMP — often blocked, but worth trying) ──
    try:
        result = subprocess.run(
            ['ping', '-c', '3', '-W', '2', host],
            capture_output=True, text=True, timeout=10
        )
        output = result.stdout.strip().split('\n')
        summary = [l for l in output if 'packet' in l.lower() or 'rtt' in l.lower() or 'avg' in l.lower()]
        add('ping', 'OK' if result.returncode == 0 else 'FAIL',
            '\n'.join(summary) if summary else result.stdout[-200:])
    except subprocess.TimeoutExpired:
        add('ping', 'TIMEOUT', 'Ping timed out — ICMP likely blocked')
    except FileNotFoundError:
        add('ping', 'SKIP', 'ping not available in container')
    except Exception as e:
        add('ping', 'ERROR', str(e))

    # ── Test 11: Tunnel URL test (if configured) ──
    if tunnel_url:
        add('tunnel_config', 'info', f'Tunnel URL configured: {tunnel_url}')
        try:
            t0 = time.time()
            r = req_lib.get(tunnel_url, headers={'ngrok-skip-browser-warning': '1'}, timeout=15)
            ms = (time.time() - t0) * 1000
            is_tally = 'Tally' in r.text
            add('tunnel_http_get', 'OK' if is_tally else 'UNEXPECTED',
                f'HTTP {r.status_code}, Tally={"yes" if is_tally else "no"}, body={r.text[:200]}', ms)
        except req_lib.exceptions.ConnectTimeout:
            add('tunnel_http_get', 'CONNECT_TIMEOUT', f'Tunnel not reachable (15s) — is ngrok running on office PC?')
        except req_lib.exceptions.ReadTimeout:
            add('tunnel_http_get', 'READ_TIMEOUT', 'Tunnel connected but no response (15s)')
        except req_lib.exceptions.ConnectionError as e:
            add('tunnel_http_get', 'CONN_ERROR', str(e)[:300])
        except Exception as e:
            add('tunnel_http_get', 'ERROR', f'{type(e).__name__}: {e}')

    # ── Summary verdict ──
    tunnel_test = next((t for t in tests if t['test'] == 'tunnel_http_get'), None)
    tcp_tally = next((t for t in tests if t['test'].startswith(f'tcp_{host}_{port}')), None)
    http_tally = next((t for t in tests if t['test'] == 'http_get_tally'), None)
    sanity = next((t for t in tests if t['test'] == 'tcp_sanity_google_443'), None)
    host_80 = next((t for t in tests if t['test'] == f'tcp_{host}_80'), None)

    if tunnel_test and tunnel_test['status'] == 'OK':
        verdict = 'TUNNEL_OK — Tally reachable via ngrok tunnel'
    elif tunnel_test and tunnel_test['status'] != 'OK':
        verdict = 'TUNNEL_FAIL — Tunnel configured but not working. Check if ngrok is running on office PC.'
    elif http_tally and http_tally['status'] == 'OK':
        verdict = 'ALL_GOOD — Tally reachable via HTTP'
    elif tcp_tally and tcp_tally['status'] == 'OK':
        verdict = 'TCP_OK_HTTP_FAIL — Port reachable but HTTP failed (Tally protocol issue?)'
    elif sanity and sanity['status'] != 'OK':
        verdict = 'NO_OUTBOUND — This server cannot make any outbound TCP connections'
    elif host_80 and host_80['status'] == 'TIMEOUT':
        verdict = 'HOST_UNREACHABLE — Cannot reach the Tally host at all (IP blocked or routing issue)'
    elif host_80 and host_80['status'] in ('OK', 'REFUSED') and tcp_tally and tcp_tally['status'] == 'TIMEOUT':
        verdict = 'PORT_BLOCKED — Host reachable on 80 but NOT on {port}. Firewall is dropping packets to port {port} from this IP.'.format(port=port)
    elif tcp_tally and tcp_tally['status'] == 'REFUSED':
        verdict = 'PORT_CLOSED — Tally not listening on port {port} (got RST)'.format(port=port)
    else:
        verdict = 'UNKNOWN — Check individual test results'

    return JsonResponse({
        'verdict': verdict,
        'target': f'{host}:{port}',
        'tests': tests,
    })


@login_required
@require_http_methods(["POST"])
def api_trigger_sync(request):
    """Trigger an incremental Tally sync now (last 7 days). Finance/Admin only."""
    if request.user.role not in ['admin', 'super_user', 'finance_manager', 'director', 'operation_controller']:
        return JsonResponse({'error': 'Access denied.'}, status=403)

    from integration_workers import create_task
    create_task(
        endpoint='/tallysync/workers/sync-tally-data/',
        payload={'sync_type': 'vouchers', 'triggered_by_user': request.user.username}
    )
    return JsonResponse({'status': 'queued', 'message': 'Incremental sync queued (last 7 days)'})


@login_required
@require_http_methods(["POST"])
def api_trigger_full_sync(request):
    """Trigger a full Tally sync from the beginning."""
    if request.user.role not in ['admin', 'super_user', 'director', 'finance_manager', 'operation_controller']:
        return JsonResponse({'error': 'Access denied.'}, status=403)

    from integration_workers import create_task
    create_task(
        endpoint='/tallysync/workers/sync-tally-data/',
        payload={'sync_type': 'all', 'triggered_by_user': request.user.username, 'full_sync': True}
    )
    return JsonResponse({'status': 'queued', 'message': 'Full sync queued — this may take several minutes'})


@login_required
@require_http_methods(["GET"])
def api_sync_logs(request):
    """Return last 20 tally batch sync logs for the sync history panel."""
    allowed_roles = ['admin', 'super_user', 'director', 'finance_manager', 'operation_controller']
    if request.user.role not in allowed_roles:
        return JsonResponse({'error': 'Access denied.'}, status=403)

    from integrations.models import SyncLog
    from django.utils import timezone as tz

    logs = SyncLog.objects.filter(
        integration='tallysync',
        log_kind='batch',
    ).order_by('-started_at')[:20]

    now = tz.now()

    def time_ago(dt):
        if not dt:
            return '—'
        diff = int((now - dt).total_seconds())
        if diff < 60:
            return f"{diff}s ago"
        if diff < 3600:
            return f"{diff // 60}m ago"
        if diff < 86400:
            return f"{diff // 3600}h ago"
        return dt.strftime('%d %b %H:%M')

    result = []
    for log in logs:
        result.append({
            'id': log.id,
            'sync_type': log.get_sync_type_display(),
            'status': log.status,
            'started_at': log.started_at.strftime('%d %b %H:%M:%S'),
            'started_ago': time_ago(log.started_at),
            'completed_at': log.completed_at.strftime('%d %b %H:%M:%S') if log.completed_at else None,
            'duration': log.duration_display,
            'records_synced': log.total_records_synced,
            'records_created': log.records_created,
            'records_updated': log.records_updated,
            'records_failed': log.records_failed,
            'triggered_by': log.triggered_by_user or log.triggered_by or 'scheduler',
            'progress': log.overall_progress_percent,
            'error': log.error_message,
        })

    return JsonResponse({'logs': result})


@login_required
@require_http_methods(["GET"])
def api_sync_log_detail(request, batch_id):
    """Return operation-level logs for a specific TallySync batch."""
    allowed_roles = ['admin', 'super_user', 'director', 'finance_manager', 'operation_controller']
    if request.user.role not in allowed_roles:
        return JsonResponse({'error': 'Access denied.'}, status=403)

    from integrations.models import SyncLog
    from django.utils import timezone as tz

    try:
        batch_log = SyncLog.objects.get(pk=batch_id, integration='tallysync', log_kind='batch')
        operation_logs = SyncLog.objects.filter(
            batch=batch_log,
            log_kind='operation'
        ).order_by('started_at')

        logs = []
        for op in operation_logs:
            logs.append({
                'id': op.id,
                'timestamp': tz.localtime(op.started_at).strftime('%H:%M:%S'),
                'level': op.level,
                'operation': op.operation,
                'message': op.message or '',
                'duration_ms': op.duration_ms,
            })

        return JsonResponse({
            'logs': logs,
            'batch_status': batch_log.status,
            'batch_started': batch_log.started_at.isoformat(),
            'batch_completed': batch_log.completed_at.isoformat() if batch_log.completed_at else None,
        })
    except SyncLog.DoesNotExist:
        return JsonResponse({'error': 'Sync log not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def api_cost_centre_profitability(request):
    """
    Cost-centre (project code) level profitability.

    Each CC maps to: code, client, vendor, location, state, company.
    Returns revenue, purchase, profit, margin aggregated per CC.
    Supports date range and company filters.
    """
    import re
    from integrations.tallysync.models import TallyVoucherCostCentreAllocation
    from django.db.models import Sum, Q, Value
    from django.db.models.functions import Coalesce

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    company_id = request.GET.get('company_id')
    search = request.GET.get('search', '').strip().lower()
    state_filter = request.GET.get('state', '').strip().upper()

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    SALES_TYPES = ['Sales']
    CN_TYPES = ['Credit Note']
    PURCHASE_TYPES = ['Purchase', 'Purchase Expenses', 'Purchase Rcm', 'Purchae Rcm']
    DN_TYPES = ['Debit Note']

    qs = TallyVoucherCostCentreAllocation.objects.exclude(
        ledger_entry__voucher__is_cancelled=True
    )
    if start_date:
        qs = qs.filter(ledger_entry__voucher__date__gte=start_date)
    if end_date:
        qs = qs.filter(ledger_entry__voucher__date__lte=end_date)
    if company_id:
        qs = qs.filter(ledger_entry__voucher__company_id=company_id)

    rows = list(
        qs.values('cost_centre_name')
        .annotate(
            gross_sales=Coalesce(Sum('amount', filter=Q(ledger_entry__voucher__voucher_type__in=SALES_TYPES)), Value(Decimal('0'))),
            credit_notes=Coalesce(Sum('amount', filter=Q(ledger_entry__voucher__voucher_type__in=CN_TYPES)), Value(Decimal('0'))),
            gross_purchase=Coalesce(Sum('amount', filter=Q(ledger_entry__voucher__voucher_type__in=PURCHASE_TYPES)), Value(Decimal('0'))),
            debit_notes=Coalesce(Sum('amount', filter=Q(ledger_entry__voucher__voucher_type__in=DN_TYPES)), Value(Decimal('0'))),
        )
    )

    # CC name pattern: "MH190 - (Client - Vendor (Location))" — location may contain nested parens e.g. "Delhi (Mundka)"
    cc_pattern = re.compile(r'^([A-Z]{1,2}[-]?\d+)\s*[-–]\s*\((.+?)\s+-\s+(.+?)\s+\((.+?)\)\s*\)?$')

    results = []
    for r in rows:
        cc = r['cost_centre_name'] or ''
        gs = r['gross_sales'] or Decimal('0')
        cn = r['credit_notes'] or Decimal('0')
        gp = r['gross_purchase'] or Decimal('0')
        dn = r['debit_notes'] or Decimal('0')

        revenue = gs - cn
        purchase = gp - dn
        profit = revenue - purchase

        # Skip CCs with no activity at all
        if revenue == 0 and purchase == 0:
            continue

        m = cc_pattern.match(cc)
        if m:
            code = m.group(1)
            client = m.group(2).strip()
            vendor = m.group(3).strip()
            location = m.group(4).strip()
            state = ''.join(filter(str.isalpha, code))
        else:
            code = cc
            client = ''
            vendor = ''
            location = ''
            state = ''

        # Apply search filter
        if search:
            haystack = f"{cc} {client} {vendor} {location}".lower()
            if search not in haystack:
                continue

        # Apply state filter
        if state_filter and state != state_filter:
            continue

        margin = float(profit / revenue * 100) if revenue > 0 else 0.0

        results.append({
            'cc_name': cc,
            'code': code,
            'client': client,
            'vendor': vendor,
            'location': location,
            'state': state,
            'revenue': revenue,
            'purchase': purchase,
            'profit': profit,
            'margin': round(margin, 1),
        })

    results.sort(key=lambda x: -x['revenue'])

    # Summary totals
    total_revenue = sum(r['revenue'] for r in results)
    total_purchase = sum(r['purchase'] for r in results)
    total_profit = sum(r['profit'] for r in results)
    total_margin = float(total_profit / total_revenue * 100) if total_revenue > 0 else 0.0

    return JsonResponse({
        'projects': results,
        'summary': {
            'total_projects': len(results),
            'total_revenue': total_revenue,
            'total_purchase': total_purchase,
            'total_profit': total_profit,
            'total_margin': round(total_margin, 1),
            'profitable': sum(1 for r in results if r['profit'] > 0),
            'loss_making': sum(1 for r in results if r['profit'] < 0),
        },
    }, encoder=DecimalEncoder)
