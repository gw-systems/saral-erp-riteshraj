"""
Google Ads Views
Dashboard, settings, OAuth, and management views
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Avg, Count, F
from django.db.models.functions import Upper
from django.urls import reverse
from datetime import datetime, timedelta
from django.utils import timezone
import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from decimal import Decimal

from .models import (
    GoogleAdsToken,
    Campaign,
    CampaignPerformance,
    DevicePerformance,
    SearchTerm,
)
from integrations.models import SyncLog
from .utils.encryption import GoogleAdsEncryption
from .utils.google_ads_auth import GoogleAdsAuth
from integration_workers import create_task

import logging
logger = logging.getLogger(__name__)


@login_required
def dashboard(request):
    """
    Main Google Ads dashboard
    Shows campaign performance overview
    """
    # RBAC: Admin/Director and CRM Executive only
    if request.user.role not in ['admin', 'director', 'crm_executive', 'digital_marketing']:
        messages.error(request, "You don't have permission to access Google Ads dashboard.")
        return redirect('accounts:home')

    # Get active tokens
    tokens = GoogleAdsToken.objects.filter(is_active=True)

    # Get filter parameters
    token_id = request.GET.get('token_id', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Default date range (current month)
    now = timezone.now()
    if not start_date:
        start_date = now.replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = now.strftime('%Y-%m-%d')

    # Build queryset for campaigns
    campaigns_qs = Campaign.objects.all()

    if token_id:
        campaigns_qs = campaigns_qs.filter(token_id=token_id)

    # Get performance data for date range
    performance_qs = CampaignPerformance.objects.all()

    if token_id:
        performance_qs = performance_qs.filter(campaign__token_id=token_id)

    if start_date:
        performance_qs = performance_qs.filter(date__gte=start_date)
    if end_date:
        performance_qs = performance_qs.filter(date__lte=end_date)

    # Calculate summary statistics
    summary_stats = performance_qs.aggregate(
        total_impressions=Sum('impressions'),
        total_clicks=Sum('clicks'),
        total_cost=Sum('cost'),
        total_conversions=Sum('conversions'),
        total_conversion_value=Sum('conversion_value'),
        avg_ctr=Avg('ctr'),
        avg_cpc=Avg('avg_cpc'),
        avg_impression_share=Avg('impression_share')
    )

    # Convert impression share to percentage (it's stored as 0-1)
    if summary_stats['avg_impression_share']:
        summary_stats['avg_impression_share'] = summary_stats['avg_impression_share'] * 100

    # Campaign performance aggregated by campaign
    campaign_performance = performance_qs.values(
        'campaign__campaign_name',
        'campaign__campaign_id',
        'campaign__campaign_status',
        'campaign__daily_budget',
        'campaign__monthly_budget',
        'campaign__bidding_strategy_type',
        'campaign__budget_delivery_method'
    ).annotate(
        impressions=Sum('impressions'),
        clicks=Sum('clicks'),
        cost=Sum('cost'),
        conversions=Sum('conversions'),
        conversion_value=Sum('conversion_value'),
        avg_ctr=Avg('ctr'),
        avg_cpc=Avg('avg_cpc'),
        avg_conversion_rate=Avg('conversion_rate'),
        avg_impression_share=Avg('impression_share') * 100,  # Convert to percentage
        avg_budget_utilization=Avg('budget_utilization') * 100  # Convert to percentage
    ).order_by('-clicks')

    # Daily performance trend (last 30 days)
    daily_trend = performance_qs.values('date').annotate(
        impressions=Sum('impressions'),
        clicks=Sum('clicks'),
        cost=Sum('cost'),
        conversions=Sum('conversions')
    ).order_by('date')

    # Device breakdown for date range - normalize device names to uppercase to avoid duplicates
    device_breakdown_raw = DevicePerformance.objects.filter(
        campaign__in=campaigns_qs,
        date__gte=start_date,
        date__lte=end_date
    ).annotate(
        device_normalized=Upper('device')
    ).values('device_normalized').annotate(
        impressions=Sum('impressions'),
        clicks=Sum('clicks'),
        cost=Sum('cost'),
        conversions=Sum('conversions')
    ).order_by('device_normalized')

    # Calculate derived metrics
    device_breakdown = []
    for device in device_breakdown_raw:
        impressions = device['impressions'] or 0
        clicks = device['clicks'] or 0
        cost = float(device['cost'] or 0)
        conversions = float(device['conversions'] or 0)

        # Rename device_normalized back to device for template
        device['device'] = device.pop('device_normalized')
        device['ctr'] = (clicks / impressions * 100) if impressions > 0 else 0
        device['avg_cpc'] = (cost / clicks) if clicks > 0 else 0
        device['conversion_rate'] = (conversions / clicks * 100) if clicks > 0 else 0
        device['cost_per_conversion'] = (cost / conversions) if conversions > 0 else 0
        device_breakdown.append(device)

    # Pagination for campaign performance
    paginator = Paginator(list(campaign_performance), 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'page_title': 'Google Ads Dashboard',
        'tokens': tokens,
        'selected_token_id': token_id,
        'start_date': start_date,
        'end_date': end_date,
        'summary_stats': summary_stats,
        'campaign_performance': page_obj,
        'daily_trend': list(daily_trend),
        'device_breakdown': device_breakdown,
        'page_obj': page_obj,
    }

    return render(request, 'google_ads/dashboard.html', context)


@login_required
def search_terms(request):
    """
    Search terms viewer
    """
    # RBAC check
    if request.user.role not in ['admin', 'director', 'crm_executive', 'digital_marketing']:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('accounts:home')

    # Get active tokens
    tokens = GoogleAdsToken.objects.filter(is_active=True)

    # Get filter parameters
    token_id = request.GET.get('token_id', '')
    year = request.GET.get('year', timezone.now().year)
    month = request.GET.get('month', timezone.now().month)
    search_query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')

    # Build queryset
    search_terms_qs = SearchTerm.objects.all()

    if token_id:
        search_terms_qs = search_terms_qs.filter(campaign__token_id=token_id)

    if year:
        search_terms_qs = search_terms_qs.filter(year=year)

    if month:
        search_terms_qs = search_terms_qs.filter(month=month)

    if search_query:
        search_terms_qs = search_terms_qs.filter(
            Q(search_term__icontains=search_query) |
            Q(campaign__campaign_name__icontains=search_query)
        )

    if status_filter:
        search_terms_qs = search_terms_qs.filter(status=status_filter)

    search_terms_qs = search_terms_qs.order_by('-clicks')

    # Summary statistics
    summary_stats = search_terms_qs.aggregate(
        total_impressions=Sum('impressions'),
        total_clicks=Sum('clicks'),
        total_cost=Sum('cost'),
        total_conversions=Sum('conversions'),
        unique_terms=Count('search_term', distinct=True)
    )

    # Pagination
    paginator = Paginator(search_terms_qs, 100)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Add formatted date to each search term for display
    from datetime import date
    for term in page_obj:
        # Create a date object from year and month (use day 1)
        term_date = date(term.year, term.month, 1)
        term.formatted_date = term_date.strftime('%d-%b-%Y')

    context = {
        'page_title': 'Search Terms',
        'tokens': tokens,
        'selected_token_id': token_id,
        'selected_year': int(year) if year else timezone.now().year,
        'selected_month': int(month) if month else timezone.now().month,
        'search_query': search_query,
        'status_filter': status_filter,
        'summary_stats': summary_stats,
        'search_terms': page_obj,
        'page_obj': page_obj,
    }

    return render(request, 'google_ads/search_terms.html', context)


@login_required
def device_performance(request):
    """
    Device performance breakdown
    """
    # RBAC check
    if request.user.role not in ['admin', 'director', 'crm_executive', 'digital_marketing']:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('accounts:home')

    # Get active tokens
    tokens = GoogleAdsToken.objects.filter(is_active=True)

    # Get filter parameters
    token_id = request.GET.get('token_id', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    device = request.GET.get('device', '')

    # Default date range (last 30 days)
    if not start_date:
        start_date = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not end_date:
        end_date = timezone.now().strftime('%Y-%m-%d')

    # Build queryset
    device_perf_qs = DevicePerformance.objects.all()

    if token_id:
        device_perf_qs = device_perf_qs.filter(campaign__token_id=token_id)

    if start_date:
        device_perf_qs = device_perf_qs.filter(date__gte=start_date)
    if end_date:
        device_perf_qs = device_perf_qs.filter(date__lte=end_date)

    if device:
        device_perf_qs = device_perf_qs.filter(device=device)

    # Device breakdown summary
    device_summary = device_perf_qs.values('device').annotate(
        impressions=Sum('impressions'),
        clicks=Sum('clicks'),
        cost=Sum('cost'),
        conversions=Sum('conversions'),
        avg_ctr=Avg('ctr'),
        avg_cpc=Avg('avg_cpc')
    ).order_by('-clicks')

    # Daily trend by device
    daily_device_trend = device_perf_qs.values('date', 'device').annotate(
        impressions=Sum('impressions'),
        clicks=Sum('clicks'),
        cost=Sum('cost')
    ).order_by('date', 'device')

    context = {
        'page_title': 'Device Performance',
        'tokens': tokens,
        'selected_token_id': token_id,
        'start_date': start_date,
        'end_date': end_date,
        'selected_device': device,
        'device_summary': list(device_summary),
        'daily_device_trend': list(daily_device_trend),
    }

    return render(request, 'google_ads/device_performance.html', context)


@login_required
def settings(request):
    """
    Google Ads settings page
    OAuth connection management (Admin/Director only)
    """
    # RBAC: Admin/Director only
    if request.user.role not in ['admin', 'director', 'digital_marketing']:
        messages.error(request, "You don't have permission to access settings.")
        return redirect('google_ads:dashboard')

    # Get all tokens
    tokens = GoogleAdsToken.objects.all().order_by('-created_at')

    context = {
        'page_title': 'Google Ads Settings',
        'tokens': tokens,
    }

    return render(request, 'google_ads/settings.html', context)


@login_required
def oauth_start(request):
    """
    Step 1: Initiate OAuth2 flow
    """
    # RBAC: Admin/Director only
    if request.user.role not in ['admin', 'director', 'digital_marketing']:
        messages.error(request, "You don't have permission to connect accounts.")
        return redirect('google_ads:dashboard')

    # Get account name and customer ID from form
    account_name = request.POST.get('account_name', '').strip()
    customer_id = request.POST.get('customer_id', '').strip()

    if not account_name or not customer_id:
        messages.error(request, "Please provide both account name and customer ID.")
        return redirect('/accounts/dashboard/admin/integrations/?tab=google_ads')

    # Store in session for callback
    request.session['google_ads_account_name'] = account_name
    request.session['google_ads_customer_id'] = customer_id

    # Use configured redirect URI (must match Google Cloud Console exactly)
    from django.conf import settings
    redirect_uri = settings.GOOGLE_ADS_REDIRECT_URI
    authorization_url, state = GoogleAdsAuth.get_authorization_url(redirect_uri)

    # Store state in session for CSRF protection
    request.session['google_ads_oauth_state'] = state

    return redirect(authorization_url)


@login_required
def oauth_callback(request):
    """
    Step 2: Handle OAuth2 callback and store tokens
    """
    # RBAC: Admin/Director only
    if request.user.role not in ['admin', 'director', 'digital_marketing']:
        messages.error(request, "You don't have permission to connect accounts.")
        return redirect('google_ads:dashboard')

    # Verify state for CSRF protection
    state = request.GET.get('state')
    session_state = request.session.get('google_ads_oauth_state')

    if not state or state != session_state:
        messages.error(request, "Invalid OAuth state. Please try again.")
        return redirect('/accounts/dashboard/admin/integrations/?tab=google_ads')

    # Get authorization code
    code = request.GET.get('code')
    if not code:
        error = request.GET.get('error', 'Unknown error')
        messages.error(request, f"OAuth authorization failed: {error}")
        return redirect('/accounts/dashboard/admin/integrations/?tab=google_ads')

    # Exchange code for token
    try:
        from django.conf import settings
        redirect_uri = settings.GOOGLE_ADS_REDIRECT_URI
        token_data = GoogleAdsAuth.exchange_code_for_token(code, redirect_uri)

        # Get account details from session
        account_name = request.session.get('google_ads_account_name', 'Unnamed Account')
        customer_id = request.session.get('google_ads_customer_id', '')

        # Encrypt and store token
        encrypted_token = GoogleAdsEncryption.encrypt(token_data)

        # Create or update token
        token, created = GoogleAdsToken.objects.update_or_create(
            customer_id=customer_id,
            defaults={
                'user': request.user,
                'account_name': account_name,
                'encrypted_token': encrypted_token,
                'is_active': True
            }
        )

        # Clean up session
        request.session.pop('google_ads_oauth_state', None)
        request.session.pop('google_ads_account_name', None)
        request.session.pop('google_ads_customer_id', None)

        if created:
            messages.success(request, f"Successfully connected Google Ads account: {account_name}")
        else:
            messages.success(request, f"Successfully updated Google Ads account: {account_name}")

        return redirect('/accounts/dashboard/admin/integrations/?tab=google_ads')

    except Exception as e:
        logger.error(f"OAuth callback error: {e}")
        messages.error(request, f"Failed to complete OAuth flow: {str(e)}")
        return redirect('/accounts/dashboard/admin/integrations/?tab=google_ads')


@login_required
def disconnect(request, token_id):
    """
    Disconnect (delete) a Google Ads account
    """
    # RBAC: Admin/Director only
    if request.user.role not in ['admin', 'director', 'digital_marketing']:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=400)

    try:
        token = get_object_or_404(GoogleAdsToken, id=token_id)
        account_name = token.account_name
        token.delete()

        return JsonResponse({
            'success': True,
            'message': f'Disconnected Google Ads account: {account_name}'
        })

    except Exception as e:
        logger.error(f"Disconnect error: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def sync_account(request, token_id):
    """
    AJAX endpoint to sync a single Google Ads account.
    Accepts optional sync_type in POST: 'campaigns', 'search_terms', or 'all' (default).
    """
    # RBAC check
    if request.user.role not in ['admin', 'director', 'crm_executive', 'digital_marketing']:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=400)

    try:
        from django.conf import settings
        token = get_object_or_404(GoogleAdsToken, id=token_id)
        sync_type = request.POST.get('sync_type', 'all')  # 'campaigns', 'search_terms', or 'all'

        # Map sync_type to sync flags
        sync_campaigns_only = sync_type == 'campaigns'
        sync_search_terms_only = sync_type == 'search_terms'

        # Create batch log before dispatching sync
        from integrations.models import SyncLog
        batch_log = SyncLog.objects.create(
            integration='google_ads',
            sync_type='google_ads',
            log_kind='batch',
            status='running',
            triggered_by_user=token.account_name,
        )

        task_name = create_task(
            endpoint='/integrations/google-ads/workers/sync-account/',
            payload={
                'token_id': token_id,
                'sync_yesterday': not sync_search_terms_only,
                'sync_current_month_search_terms': not sync_campaigns_only,
                'triggered_by_user': request.user.username,
                'batch_log_id': batch_log.id,
            },
            task_name=f'google-ads-sync-{token_id}-{int(timezone.now().timestamp())}'
        )

        return JsonResponse({
            'success': True,
            'message': f'Sync started for {token.account_name}',
            'task_name': task_name,
            'sync_id': batch_log.id
        })

    except Exception as e:
        logger.error(f"Sync error: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def sync_all_accounts(request):
    """
    AJAX endpoint to sync all active Google Ads accounts
    """
    # RBAC check
    if request.user.role not in ['admin', 'director', 'digital_marketing']:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=400)

    try:
        active_tokens = GoogleAdsToken.objects.filter(is_active=True)

        if not active_tokens.exists():
            return JsonResponse({
                'success': False,
                'error': 'No active Google Ads accounts found'
            })

        # Trigger Cloud Tasks worker for all accounts
        task_name = create_task(
            endpoint='/integrations/google-ads/workers/sync-all-accounts/',
            payload={
                'sync_yesterday': True,
                'sync_current_month_search_terms': True
            },
            task_name=f'google-ads-sync-all-{int(timezone.now().timestamp())}'
        )

        return JsonResponse({
            'success': True,
            'message': f'Sync started for {active_tokens.count()} accounts',
            'task_name': task_name
        })

    except Exception as e:
        logger.error(f"Sync all error: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def sync_historical(request, token_id):
    """
    AJAX endpoint to sync all historical data (everything available)
    """
    # RBAC: Admin/Director/Digital Marketing only
    if request.user.role not in ['admin', 'director', 'digital_marketing']:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=400)

    try:
        from django.conf import settings
        token = get_object_or_404(GoogleAdsToken, id=token_id)
        # Use earliest possible date to fetch everything (2018-01-01)
        start_date = request.POST.get('start_date', '2018-01-01')

        # Create batch log before dispatching sync
        from integrations.models import SyncLog
        batch_log = SyncLog.objects.create(
            integration='google_ads',
            sync_type='google_ads_historical',
            log_kind='batch',
            status='running',
            triggered_by_user=token.account_name,
        )

        task_name = create_task(
            endpoint='/integrations/google-ads/workers/sync-historical/',
            payload={
                'token_id': token_id,
                'start_date': start_date,
                'triggered_by_user': request.user.username,
                'batch_log_id': batch_log.id,
            },
            task_name=f'google-ads-historical-{token_id}-{int(timezone.now().timestamp())}',
            timeout=1800  # 30 minutes for historical sync
        )

        return JsonResponse({
            'success': True,
            'message': f'Full sync started for {token.account_name} (all available data)',
            'task_name': task_name,
            'sync_id': batch_log.id
        })

    except Exception as e:
        logger.error(f"Historical sync error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def sync_progress(request, token_id):
    """
    AJAX endpoint to get current sync progress
    Frontend polls this every 2 seconds during sync
    """
    from .sync_progress import get_sync_progress

    progress = get_sync_progress(token_id)

    if not progress:
        return JsonResponse({
            'status': 'no_sync',
            'message': 'No active sync found'
        })

    return JsonResponse(progress)


@login_required
def sync_date_range(request, token_id):
    """
    AJAX endpoint to sync specific date range
    """
    # RBAC: Admin/Director only
    if request.user.role not in ['admin', 'director', 'crm_executive', 'digital_marketing']:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=400)

    try:
        from .google_ads_sync import GoogleAdsSync

        token = get_object_or_404(GoogleAdsToken, id=token_id)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        if not start_date or not end_date:
            return JsonResponse({'success': False, 'error': 'Both start_date and end_date required'}, status=400)

        # Run sync directly (not via Celery for custom date range)
        sync_engine = GoogleAdsSync(token_id)

        # Sync campaigns first
        sync_engine.sync_campaigns()

        # Sync performance data for date range
        perf_stats = sync_engine.sync_campaign_performance(start_date, end_date)
        device_stats = sync_engine.sync_device_performance(start_date, end_date)

        # Sync search terms for months in range
        from datetime import datetime
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')

        search_terms_total = {'created': 0, 'updated': 0, 'total': 0}
        current = start
        while current <= end:
            st_stats = sync_engine.sync_search_terms(current.year, current.month)
            search_terms_total['created'] += st_stats['created']
            search_terms_total['updated'] += st_stats['updated']
            search_terms_total['total'] += st_stats['total']

            # Move to next month
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

        return JsonResponse({
            'success': True,
            'message': f'Sync completed for {token.account_name} ({start_date} to {end_date})',
            'stats': {
                'performance': perf_stats,
                'device_performance': device_stats,
                'search_terms': search_terms_total
            }
        })

    except Exception as e:
        logger.error(f"Date range sync error: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def sync_logs(request):
    """
    Sync logs viewer
    """
    # RBAC check
    if request.user.role not in ['admin', 'director', 'crm_executive', 'digital_marketing']:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('accounts:home')

    # Get filter parameters
    token_id = request.GET.get('token_id', '')
    level = request.GET.get('level', '')

    # Build queryset using unified SyncLog
    batch_logs = SyncLog.objects.filter(integration='google_ads', log_kind='batch').order_by('-started_at')[:50]
    logs_qs = SyncLog.objects.filter(integration='google_ads', log_kind='operation').order_by('-started_at')

    if level:
        logs_qs = logs_qs.filter(level=level)

    # Limit to last 1000 logs
    logs_qs = logs_qs[:1000]

    # Get active tokens for filter
    tokens = GoogleAdsToken.objects.filter(is_active=True)

    # Pagination
    paginator = Paginator(logs_qs, 100)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'page_title': 'Sync Logs',
        'tokens': tokens,
        'selected_token_id': token_id,
        'selected_level': level,
        'batch_logs': batch_logs,
        'logs': page_obj,
        'page_obj': page_obj,
    }

    return render(request, 'google_ads/sync_logs.html', context)


@login_required
def export_data(request):
    """
    Export campaign performance data to Excel or CSV
    """
    # RBAC check
    if request.user.role not in ['admin', 'director', 'crm_executive', 'digital_marketing']:
        messages.error(request, "You don't have permission to export data.")
        return redirect('accounts:home')

    # Get filter parameters
    token_id = request.GET.get('token_id', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    format_type = request.GET.get('format', 'xlsx')  # xlsx or csv

    # Default date range (last 30 days)
    if not start_date:
        start_date = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not end_date:
        end_date = timezone.now().strftime('%Y-%m-%d')

    # Build queryset
    performance_qs = CampaignPerformance.objects.all()

    if token_id:
        performance_qs = performance_qs.filter(campaign__token_id=token_id)

    if start_date:
        performance_qs = performance_qs.filter(date__gte=start_date)
    if end_date:
        performance_qs = performance_qs.filter(date__lte=end_date)

    performance_qs = performance_qs.select_related('campaign').order_by('date', 'campaign__campaign_name')

    if format_type == 'csv':
        # CSV Export
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="google_ads_performance_{start_date}_to_{end_date}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Date', 'Campaign Name', 'Campaign ID', 'Status',
            'Impressions', 'Clicks', 'Cost', 'Conversions', 'Conversion Value',
            'CTR (%)', 'Avg CPC', 'Avg CPM', 'Conversion Rate (%)', 'Cost/Conversion'
        ])

        for perf in performance_qs:
            writer.writerow([
                perf.date,
                perf.campaign.campaign_name,
                perf.campaign.campaign_id,
                perf.campaign.campaign_status,
                perf.impressions,
                perf.clicks,
                float(perf.cost),
                float(perf.conversions),
                float(perf.conversion_value),
                float(perf.ctr * 100),
                float(perf.avg_cpc),
                float(perf.avg_cpm),
                float(perf.conversion_rate * 100),
                float(perf.cost_per_conversion)
            ])

        return response

    else:
        # Excel Export
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Campaign Performance"

        # Header row
        headers = [
            'Date', 'Campaign Name', 'Campaign ID', 'Status',
            'Impressions', 'Clicks', 'Cost', 'Conversions', 'Conversion Value',
            'CTR (%)', 'Avg CPC', 'Avg CPM', 'Conversion Rate (%)', 'Cost/Conversion'
        ]

        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        for row_num, perf in enumerate(performance_qs, 2):
            ws.cell(row=row_num, column=1, value=perf.date)
            ws.cell(row=row_num, column=2, value=perf.campaign.campaign_name)
            ws.cell(row=row_num, column=3, value=perf.campaign.campaign_id)
            ws.cell(row=row_num, column=4, value=perf.campaign.campaign_status)
            ws.cell(row=row_num, column=5, value=perf.impressions)
            ws.cell(row=row_num, column=6, value=perf.clicks)
            ws.cell(row=row_num, column=7, value=float(perf.cost))
            ws.cell(row=row_num, column=8, value=float(perf.conversions))
            ws.cell(row=row_num, column=9, value=float(perf.conversion_value))
            ws.cell(row=row_num, column=10, value=float(perf.ctr * 100))
            ws.cell(row=row_num, column=11, value=float(perf.avg_cpc))
            ws.cell(row=row_num, column=12, value=float(perf.avg_cpm))
            ws.cell(row=row_num, column=13, value=float(perf.conversion_rate * 100))
            ws.cell(row=row_num, column=14, value=float(perf.cost_per_conversion))

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="google_ads_performance_{start_date}_to_{end_date}.xlsx"'

        wb.save(response)
        return response


@login_required
def detailed_report(request):
    """
    Detailed performance report with device breakdowns
    Matches the exact format: Date | Campaign | Cost | Device metrics
    """
    # RBAC: Admin/Director and Digital Marketing only
    if request.user.role not in ['admin', 'director', 'crm_executive', 'digital_marketing']:
        messages.error(request, "You don't have permission to access detailed reports.")
        return redirect('accounts:home')

    # Get filter parameters
    campaign_name = request.GET.get('campaign_name', '')
    campaign_status = request.GET.get('campaign_status', '')

    # Default date range (current month)
    now = timezone.now()
    start_date = request.GET.get('start_date', now.replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', now.strftime('%Y-%m-%d'))
    export_format = request.GET.get('export', '')

    # Base querysets
    performance_qs = CampaignPerformance.objects.select_related('campaign', 'campaign__token').all()
    device_qs = DevicePerformance.objects.select_related('campaign', 'campaign__token').all()

    # Apply filters
    if campaign_name:
        performance_qs = performance_qs.filter(campaign__campaign_name__icontains=campaign_name)
        device_qs = device_qs.filter(campaign__campaign_name__icontains=campaign_name)

    if campaign_status:
        performance_qs = performance_qs.filter(campaign__campaign_status=campaign_status)
        device_qs = device_qs.filter(campaign__campaign_status=campaign_status)

    if start_date:
        performance_qs = performance_qs.filter(date__gte=start_date)
        device_qs = device_qs.filter(date__gte=start_date)

    if end_date:
        performance_qs = performance_qs.filter(date__lte=end_date)
        device_qs = device_qs.filter(date__lte=end_date)

    # Get unique campaign names and statuses for dropdowns
    all_campaigns = Campaign.objects.filter(token__is_active=True).values_list('campaign_name', flat=True).distinct().order_by('campaign_name')
    all_statuses = Campaign.objects.filter(token__is_active=True).values_list('campaign_status', flat=True).distinct().order_by('campaign_status')

    # Build detailed report data by merging performance + device data
    report_data = []
    total_cost = Decimal('0')
    total_clicks = 0
    total_impressions = Decimal('0')
    total_conversions = Decimal('0')
    ctr_sum = Decimal('0')
    ctr_count = 0

    for perf in performance_qs.order_by('-date', 'campaign__campaign_name'):
        # Get device performance for this campaign and date
        devices = device_qs.filter(campaign=perf.campaign, date=perf.date)

        mobile = devices.filter(device='MOBILE').first()
        desktop = devices.filter(device='COMPUTER').first()
        tablet = devices.filter(device='TABLET').first()

        # Calculate CTR
        ctr = (float(perf.ctr) * 100) if perf.ctr else 0
        # Calculate conversion rate
        conv_rate = (float(perf.conversions) / perf.clicks * 100) if perf.clicks > 0 else 0
        # Calculate cost per conversion
        cost_per_conv = (float(perf.cost) / float(perf.conversions)) if perf.conversions > 0 else 0

        # Monthly budget (assuming 28 days on average)
        monthly_budget = float(perf.campaign.daily_budget) * 28 if perf.campaign.daily_budget else 0

        row = {
            'date': perf.date.strftime('%d-%b-%Y'),
            'campaign_id': perf.campaign.campaign_id,
            'campaign_name': perf.campaign.campaign_name,
            'campaign_status': perf.campaign.campaign_status,
            'daily_budget': float(perf.campaign.daily_budget) if perf.campaign.daily_budget else 0,
            'monthly_budget': monthly_budget,
            'cost': float(perf.cost),
            'impression_share': float(perf.impression_share) * 100 if perf.impression_share else 0,
            'impressions': perf.impressions,
            'clicks': perf.clicks,
            'ctr': ctr,
            'cpc': float(perf.avg_cpc) if perf.avg_cpc else 0,
            'conversions': int(perf.conversions),
            'conversion_rate': conv_rate,
            'cost_per_conversion': cost_per_conv,
            # Mobile
            'mobile_clicks': mobile.clicks if mobile else 0,
            'mobile_impressions': mobile.impressions if mobile else 0,
            'mobile_cost': float(mobile.cost) if mobile and mobile.cost else 0,
            'mobile_conversions': int(mobile.conversions) if mobile and mobile.conversions else 0,
            # Desktop
            'desktop_clicks': desktop.clicks if desktop else 0,
            'desktop_impressions': desktop.impressions if desktop else 0,
            'desktop_cost': float(desktop.cost) if desktop and desktop.cost else 0,
            'desktop_conversions': int(desktop.conversions) if desktop and desktop.conversions else 0,
            # Tablet
            'tablet_clicks': tablet.clicks if tablet else 0,
            'tablet_impressions': tablet.impressions if tablet else 0,
            'tablet_cost': float(tablet.cost) if tablet and tablet.cost else 0,
            'tablet_conversions': int(tablet.conversions) if tablet and tablet.conversions else 0,
        }

        report_data.append(row)

        # Update totals
        total_cost += perf.cost
        total_clicks += perf.clicks
        total_impressions += perf.impressions
        total_conversions += perf.conversions
        if perf.ctr:
            ctr_sum += Decimal(str(perf.ctr)) * 100
            ctr_count += 1

    avg_ctr = float(ctr_sum / ctr_count) if ctr_count > 0 else 0

    # Handle exports
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="google_ads_detailed_{start_date}_to_{end_date}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Date', 'Campaign ID', 'Campaign Name', 'Campaign Status',
            'Daily Budget', 'Monthly Budget', 'Cost', 'Impression Share', 'Impressions',
            'Clicks', 'CTR %', 'CPC', 'Conversions', 'Conversion Rate', 'Cost Per Conversion',
            'Mobile Clicks', 'Mobile Impressions', 'Mobile Cost', 'Mobile Conversions',
            'Desktop Clicks', 'Desktop Impressions', 'Desktop Cost', 'Desktop Conversions',
            'Tablet Clicks', 'Tablet Impressions', 'Tablet Cost', 'Tablet Conversions'
        ])

        for row in report_data:
            writer.writerow([
                row['date'], row['campaign_id'], row['campaign_name'], row['campaign_status'],
                f"₹{row['daily_budget']:.2f}", f"₹{row['monthly_budget']:.2f}",
                f"₹{row['cost']:.2f}", f"{row['impression_share']:.2f}%", row['impressions'],
                row['clicks'], f"{row['ctr']:.2f}%", f"₹{row['cpc']:.2f}",
                row['conversions'], f"{row['conversion_rate']:.2f}%", f"₹{row['cost_per_conversion']:.2f}",
                row['mobile_clicks'], row['mobile_impressions'], f"₹{row['mobile_cost']:.2f}", row['mobile_conversions'],
                row['desktop_clicks'], row['desktop_impressions'], f"₹{row['desktop_cost']:.2f}", row['desktop_conversions'],
                row['tablet_clicks'], row['tablet_impressions'], f"₹{row['tablet_cost']:.2f}", row['tablet_conversions']
            ])

        return response

    elif export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detailed Performance"

        # Headers
        headers = [
            'Date', 'Campaign ID', 'Campaign Name', 'Campaign Status',
            'Daily Budget', 'Monthly Budget', 'Cost', 'Impression Share', 'Impressions',
            'Clicks', 'CTR %', 'CPC', 'Conversions', 'Conversion Rate', 'Cost Per Conversion',
            'Mobile Clicks', 'Mobile Impressions', 'Mobile Cost', 'Mobile Conversions',
            'Desktop Clicks', 'Desktop Impressions', 'Desktop Cost', 'Desktop Conversions',
            'Tablet Clicks', 'Tablet Impressions', 'Tablet Cost', 'Tablet Conversions'
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        for row_num, row in enumerate(report_data, 2):
            ws.cell(row=row_num, column=1, value=row['date'])
            ws.cell(row=row_num, column=2, value=row['campaign_id'])
            ws.cell(row=row_num, column=3, value=row['campaign_name'])
            ws.cell(row=row_num, column=4, value=row['campaign_status'])
            ws.cell(row=row_num, column=5, value=row['daily_budget'])
            ws.cell(row=row_num, column=6, value=row['monthly_budget'])
            ws.cell(row=row_num, column=7, value=row['cost'])
            ws.cell(row=row_num, column=8, value=row['impression_share'])
            ws.cell(row=row_num, column=9, value=row['impressions'])
            ws.cell(row=row_num, column=10, value=row['clicks'])
            ws.cell(row=row_num, column=11, value=row['ctr'])
            ws.cell(row=row_num, column=12, value=row['cpc'])
            ws.cell(row=row_num, column=13, value=row['conversions'])
            ws.cell(row=row_num, column=14, value=row['conversion_rate'])
            ws.cell(row=row_num, column=15, value=row['cost_per_conversion'])
            ws.cell(row=row_num, column=16, value=row['mobile_clicks'])
            ws.cell(row=row_num, column=17, value=row['mobile_impressions'])
            ws.cell(row=row_num, column=18, value=row['mobile_cost'])
            ws.cell(row=row_num, column=19, value=row['mobile_conversions'])
            ws.cell(row=row_num, column=20, value=row['desktop_clicks'])
            ws.cell(row=row_num, column=21, value=row['desktop_impressions'])
            ws.cell(row=row_num, column=22, value=row['desktop_cost'])
            ws.cell(row=row_num, column=23, value=row['desktop_conversions'])
            ws.cell(row=row_num, column=24, value=row['tablet_clicks'])
            ws.cell(row=row_num, column=25, value=row['tablet_impressions'])
            ws.cell(row=row_num, column=26, value=row['tablet_cost'])
            ws.cell(row=row_num, column=27, value=row['tablet_conversions'])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="google_ads_detailed_{start_date}_to_{end_date}.xlsx"'

        wb.save(response)
        return response

    # Regular HTML view
    context = {
        'page_title': 'Detailed Performance Report',
        'all_campaigns': all_campaigns,
        'all_statuses': all_statuses,
        'selected_campaign': campaign_name,
        'selected_status': campaign_status,
        'report_data': report_data,
        'start_date': start_date,
        'end_date': end_date,
        'total_cost': total_cost,
        'total_clicks': total_clicks,
        'total_impressions': total_impressions,
        'total_conversions': total_conversions,
        'avg_ctr': avg_ctr,
    }

    return render(request, 'google_ads/detailed_report.html', context)


# ─── Stop / Force-Stop Sync ───────────────────────────────────────────────────

@login_required
def stop_sync(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    if request.user.role not in ['admin', 'director']:
        return JsonResponse({'error': 'Access denied'}, status=403)
    sync_id = request.POST.get('sync_id')
    if not sync_id:
        return JsonResponse({'error': 'sync_id is required'}, status=400)
    try:
        sync_log = SyncLog.objects.get(id=sync_id, integration='google_ads', log_kind='batch')
    except SyncLog.DoesNotExist:
        return JsonResponse({'error': f'Sync {sync_id} not found'}, status=404)
    if sync_log.status != 'running':
        return JsonResponse({'error': f'Sync is not running (status: {sync_log.status})'}, status=400)
    sync_log.stop_requested = True
    sync_log.save(update_fields=['stop_requested'])
    logger.info(f"[Google Ads] Stop requested for sync {sync_id} by {request.user}")
    return JsonResponse({'status': 'success', 'message': 'Stop requested. Sync will finish current phase then stop.'})


@login_required
def force_stop_sync(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    if request.user.role not in ['admin', 'director']:
        return JsonResponse({'error': 'Access denied'}, status=403)
    sync_id = request.POST.get('sync_id')
    if not sync_id:
        return JsonResponse({'error': 'sync_id is required'}, status=400)
    try:
        if sync_id:
            sync_log = SyncLog.objects.get(id=sync_id, integration='google_ads', log_kind='batch')
        else:
            sync_log = SyncLog.objects.filter(
                integration='google_ads', log_kind='batch'
            ).order_by('-started_at').first()
            if not sync_log:
                return JsonResponse({'error': 'No sync log found'}, status=404)
    except SyncLog.DoesNotExist:
        return JsonResponse({'error': f'Sync {sync_id} not found'}, status=404)
    elapsed = int((timezone.now() - sync_log.started_at).total_seconds())
    sync_log.status = 'stopped'
    sync_log.stop_requested = True
    sync_log.completed_at = timezone.now()
    sync_log.duration_seconds = elapsed
    sync_log.error_message = f'Force-stopped by {request.user} after {elapsed}s'
    sync_log.save()
    logger.warning(f"[Google Ads] Force-stopped sync {sync_log.id} by {request.user} after {elapsed}s")
    return JsonResponse({'status': 'success', 'message': f'Sync force-stopped after {elapsed}s.'})


@login_required
def api_sync_logs(request, batch_id):
    """
    API endpoint to fetch detailed operation logs for a specific sync batch.

    Args:
        batch_id: SyncLog batch ID

    Returns:
        JSON with operation-level logs
    """
    try:
        # Get the batch log
        batch_log = SyncLog.objects.get(pk=batch_id, integration='google_ads', log_kind='batch')

        # Get all operation logs for this batch
        operation_logs = SyncLog.objects.filter(
            batch=batch_log,
            log_kind='operation'
        ).order_by('started_at')

        # Format logs for frontend
        logs = []
        for op_log in operation_logs:
            logs.append({
                'id': op_log.id,
                'timestamp': timezone.localtime(op_log.started_at).strftime('%H:%M:%S'),
                'level': op_log.level,
                'operation': op_log.operation,
                'message': op_log.message or '',
                'duration_ms': op_log.duration_ms
            })

        return JsonResponse({
            'logs': logs,
            'batch_status': batch_log.status,
            'batch_started': batch_log.started_at.isoformat(),
            'batch_completed': batch_log.completed_at.isoformat() if batch_log.completed_at else None
        })

    except SyncLog.DoesNotExist:
        return JsonResponse({'error': 'Sync log not found'}, status=404)
    except Exception as e:
        logger.error(f"Failed to fetch sync logs: {e}")
        return JsonResponse({'error': str(e)}, status=500)
