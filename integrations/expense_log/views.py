"""
Expense Log views - OAuth flow, sync operations, dashboard.
"""
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.urls import reverse
from django.utils import timezone
try:
    from google.cloud import tasks_v2
except ImportError:
    tasks_v2 = None  # Cloud Tasks optional for local development
import io
import json
import logging
import os
import re
from collections import defaultdict
from datetime import datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.db.models import Q, Sum
from django.db.models.functions import Coalesce

from .models import (
    ExpenseLogSettings,
    GoogleSheetsToken,
    ExpenseRecord,
    UserNameMapping
)
from .utils.sheets_auth import SheetsOAuthManager
from .utils.encryption import ExpenseLogEncryption
from .utils.settings_helper import ExpenseLogSettingsHelper
from .expense_log_sync import ExpenseLogSyncEngine

logger = logging.getLogger(__name__)


@login_required
def settings_view(request):
    """
    Expense Log settings page - OAuth management, sync controls.
    """
    if request.method == 'POST':
        action = request.POST.get('action')

        # Save OAuth credentials
        if action == 'save_oauth_credentials':
            settings_obj = ExpenseLogSettings.load()
            settings_obj.client_id = request.POST.get('client_id', '').strip()

            client_secret = request.POST.get('client_secret', '').strip()
            if client_secret:
                settings_obj.encrypted_client_secret = ExpenseLogEncryption.encrypt(client_secret)

            settings_obj.redirect_uri = request.POST.get('redirect_uri', '').strip()
            settings_obj.api_version = request.POST.get('api_version', 'v4').strip()
            settings_obj.updated_by = request.user
            settings_obj.save()

            messages.success(request, "OAuth credentials saved successfully")
            return redirect('expense_log:settings')

        # Connect new Google account
        elif action == 'connect_oauth':
            sheet_id = request.POST.get('sheet_id', '').strip()
            sheet_name = request.POST.get('sheet_name', 'Sheet1').strip()

            if not sheet_id:
                messages.error(request, "Sheet ID is required")
                return redirect('expense_log:settings')

            # Store sheet info in session for callback
            request.session['expense_log_sheet_id'] = sheet_id
            request.session['expense_log_sheet_name'] = sheet_name

            # Generate OAuth URL
            try:
                oauth_manager = SheetsOAuthManager()
                auth_url, state = oauth_manager.get_authorization_url()
                request.session['expense_log_oauth_state'] = state
                return redirect(auth_url)
            except Exception as e:
                messages.error(request, f"OAuth error: {str(e)}")
                return redirect('expense_log:settings')

        # Disconnect account
        elif action == 'disconnect':
            token_id = request.POST.get('token_id')
            try:
                token = GoogleSheetsToken.objects.get(pk=token_id, user=request.user)
                token.delete()
                messages.success(request, "Account disconnected successfully")
            except GoogleSheetsToken.DoesNotExist:
                messages.error(request, "Token not found")
            return redirect('expense_log:settings')

        # Trigger sync
        elif action in ['sync_now', 'sync_full']:
            token_id = request.POST.get('token_id')
            sync_type = 'full' if action == 'sync_full' else 'incremental'

            try:
                token = GoogleSheetsToken.objects.get(pk=token_id, user=request.user)
                _trigger_sync_task(token.id, sync_type, triggered_by_user=request.user.username)
                messages.success(request, f"{sync_type.capitalize()} sync started")
            except GoogleSheetsToken.DoesNotExist:
                messages.error(request, "Token not found")
            return redirect('expense_log:settings')

    # GET request - display settings
    context = {
        'settings': ExpenseLogSettings.load(),
        'connected_accounts': GoogleSheetsToken.objects.filter(
            user=request.user,
            is_active=True
        ).order_by('-created_at'),
        'oauth_config': ExpenseLogSettingsHelper.get_oauth_config(),
    }

    return render(request, 'expense_log/settings.html', context)


@login_required
def oauth_callback(request):
    """
    OAuth callback handler - exchanges code for token.
    """
    # Get authorization response
    authorization_response = request.build_absolute_uri()
    state = request.session.get('expense_log_oauth_state')

    try:
        oauth_manager = SheetsOAuthManager()
        token_data = oauth_manager.exchange_code_for_token(authorization_response, state)

        # Encrypt token
        encrypted_token = ExpenseLogEncryption.encrypt(token_data['token'])

        # Get sheet info from session
        sheet_id = request.session.get('expense_log_sheet_id', '')
        sheet_name = request.session.get('expense_log_sheet_name', 'Sheet1')

        # Create token record
        GoogleSheetsToken.objects.update_or_create(
            user=request.user,
            email_account=token_data['email'],
            sheet_id=sheet_id,
            defaults={
                'encrypted_token': encrypted_token,
                'sheet_name': sheet_name,
                'is_active': True,
            }
        )

        # Clean up session
        request.session.pop('expense_log_oauth_state', None)
        request.session.pop('expense_log_sheet_id', None)
        request.session.pop('expense_log_sheet_name', None)

        messages.success(request, f"Successfully connected {token_data['email']}")

    except Exception as e:
        logger.error(f"OAuth callback error: {e}", exc_info=True)
        messages.error(request, f"OAuth failed: {str(e)}")

    # Redirect to integrations hub Expense Log tab
    return redirect('/accounts/dashboard/admin/integrations/?tab=expense_log')


@login_required
def dashboard(request):
    """
    Expense Log dashboard - view synced expenses.

    USER PERMISSION LOGIC:
    ----------------------
    1. Admin/Director/Accounts Executive: See ALL expenses
    2. Regular users: See only expenses where submitted_by matches their UserNameMapping.sheet_name
    3. Users without mapping: See NO expenses

    FILTERS AVAILABLE:
    ------------------
    - Universal Search: Searches across expense #, client, nature, submitted by, remark, etc.
    - Status: Approved/Pending/Rejected
    - Nature of Expense: Transport/Operation/Stationary/Other
    - Service Month: Filter by month (e.g., "February 2026")
    """
    # Get user's expenses (filtered by role and mapping)
    expenses = ExpenseRecord.get_expenses_for_user(request.user).select_related(
        'token'
    ).order_by('-timestamp')

    # Filters - default to current month
    from datetime import datetime
    current_month = datetime.now().strftime('%B %Y')  # e.g., "February 2026"
    status_filter = request.GET.get('status', '')
    month_filter = request.GET.get('month', current_month)
    nature_filter = request.GET.get('nature', '')
    search_query = request.GET.get('search', '').strip()

    if status_filter:
        expenses = expenses.filter(approval_status=status_filter)

    if month_filter:
        expenses = expenses.filter(service_month=month_filter)

    if nature_filter:
        expenses = expenses.filter(nature_of_expense=nature_filter)

    if search_query:
        # Universal search across multiple fields
        from django.db.models import Q
        expenses = expenses.filter(
            Q(unique_expense_number__icontains=search_query) |
            Q(client_name__icontains=search_query) |
            Q(client__icontains=search_query) |
            Q(submitted_by__icontains=search_query) |
            Q(nature_of_expense__icontains=search_query) |
            Q(payment_method__icontains=search_query) |
            Q(expenses_borne_by__icontains=search_query) |
            Q(remark__icontains=search_query)
        )

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(expenses, 50)  # 50 per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Summary stats
    from decimal import Decimal
    total_count_val = expenses.count()
    total_amount = sum(e.amount or Decimal('0') for e in expenses)

    # Approved stats
    approved_expenses = expenses.filter(approval_status='Approved')
    approved_count = approved_expenses.count()
    approved_amount = sum(e.amount or Decimal('0') for e in approved_expenses)

    # Pending stats
    pending_expenses = expenses.filter(approval_status='Pending')
    pending_count = pending_expenses.count()
    pending_amount = sum(e.amount or Decimal('0') for e in pending_expenses)

    # Rejected stats
    rejected_expenses = expenses.filter(approval_status='Rejected')
    rejected_count = rejected_expenses.count()
    rejected_amount = sum(e.amount or Decimal('0') for e in rejected_expenses)

    # Get distinct natures for filter dropdown
    distinct_natures = ExpenseRecord.get_expenses_for_user(request.user).exclude(
        nature_of_expense__isnull=True
    ).exclude(
        nature_of_expense=''
    ).values_list('nature_of_expense', flat=True).distinct().order_by('nature_of_expense')

    # Get distinct service months for filter dropdown
    distinct_months = ExpenseRecord.get_expenses_for_user(request.user).exclude(
        service_month__isnull=True
    ).exclude(
        service_month=''
    ).values_list('service_month', flat=True).distinct().order_by('-service_month')

    # Check user's mapping status for helpful empty state messages
    user_mapping = None
    try:
        user_mapping = UserNameMapping.objects.get(erp_user=request.user)
    except UserNameMapping.DoesNotExist:
        pass

    context = {
        'page_obj': page_obj,
        'total_count': total_count_val,
        'total_amount': total_amount,
        'approved_count': approved_count,
        'approved_amount': approved_amount,
        'pending_count': pending_count,
        'pending_amount': pending_amount,
        'rejected_count': rejected_count,
        'rejected_amount': rejected_amount,
        'status_filter': status_filter,
        'month_filter': month_filter,
        'nature_filter': nature_filter,
        'search_query': search_query,
        'distinct_natures': distinct_natures,
        'distinct_months': distinct_months,
        'user_mapping': user_mapping,
        'current_month': current_month,
    }

    return render(request, 'expense_log/dashboard.html', context)


@login_required
@require_http_methods(['GET'])
def sync_progress(request):
    """
    API endpoint for sync progress polling.

    Query params:
        token_id: GoogleSheetsToken ID

    Returns:
        JSON with progress data
    """
    token_id = request.GET.get('token_id')
    if not token_id:
        return JsonResponse({'error': 'token_id required'}, status=400)

    # Verify user owns this token
    try:
        token = GoogleSheetsToken.objects.get(pk=token_id, user=request.user)
    except GoogleSheetsToken.DoesNotExist:
        return JsonResponse({'error': 'Token not found'}, status=404)

    # Get progress from cache
    progress_data = ExpenseLogSyncEngine.get_progress(token_id)

    if not progress_data:
        return JsonResponse({
            'status': 'idle',
            'progress_percentage': 0,
            'message': 'No active sync',
            'server_logs': [],
            'can_start': True,
            'can_stop': False,
        })

    return JsonResponse(progress_data)


@login_required
def sync_logs(request, batch_id):
    """
    API endpoint to fetch detailed operation logs for a specific sync batch.

    Args:
        batch_id: SyncLog batch ID

    Returns:
        JSON with operation-level logs
    """
    from integrations.models import SyncLog

    try:
        # Get the batch log
        batch_log = SyncLog.objects.get(pk=batch_id, integration='expense_log', log_kind='batch')

        # Get all operation logs for this batch
        operation_logs = SyncLog.objects.filter(
            batch=batch_log,
            log_kind='operation'
        ).order_by('started_at')

        # Format logs for frontend
        logs = []
        for op_log in operation_logs:
            logs.append({
                'id': op_log.id,
                'timestamp': timezone.localtime(op_log.started_at).strftime('%H:%M:%S'),
                'level': op_log.level,
                'operation': op_log.operation,
                'message': op_log.message or '',
                'duration_ms': op_log.duration_ms
            })

        return JsonResponse({
            'logs': logs,
            'batch_status': batch_log.status,
            'batch_started': batch_log.started_at.isoformat(),
            'batch_completed': batch_log.completed_at.isoformat() if batch_log.completed_at else None
        })

    except SyncLog.DoesNotExist:
        return JsonResponse({'error': 'Sync log not found'}, status=404)
    except Exception as e:
        logger.error(f"Failed to fetch sync logs: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def expense_detail_api(request, expense_id):
    """
    API endpoint to fetch detailed information for a single expense.

    Args:
        expense_id: ExpenseRecord primary key

    Returns:
        JSON with expense details
    """
    try:
        # Get the expense
        expense = ExpenseRecord.objects.get(pk=expense_id)

        # Check permissions - users can only see their own expenses
        if not request.user.role in ['admin', 'director', 'accounts_executive']:
            try:
                mapping = UserNameMapping.objects.get(erp_user=request.user)
                # Support both old CharField (string) and new JSONField (list)
                if isinstance(mapping.sheet_name, list):
                    # Check if "ALL_NAMES" special value is present
                    if "ALL_NAMES" not in mapping.sheet_name:
                        # Check if expense submitted_by is in the user's mapped names
                        if expense.submitted_by not in mapping.sheet_name:
                            return JsonResponse({'error': 'Permission denied'}, status=403)
                else:
                    # Old CharField format
                    if expense.submitted_by != mapping.sheet_name:
                        return JsonResponse({'error': 'Permission denied'}, status=403)
            except UserNameMapping.DoesNotExist:
                return JsonResponse({'error': 'Permission denied'}, status=403)

        # Format amount with proper formatting
        _amount = expense.amount if expense.amount is not None else expense.charges_at_gw
        amount_display = f"{_amount:,.2f}" if _amount else "0.00"

        # Format timestamp
        timestamp_display = expense.timestamp.strftime('%b %d, %Y %I:%M %p') if expense.timestamp else '-'

        # Extract attachments from all upload invoice fields
        attachments = []
        # Transport invoices
        if expense.transport_bill:
            attachments.append({'label': 'Transport Bill', 'url': expense.transport_bill})
        if expense.upload_invoice_transport_2:
            attachments.append({'label': 'Transport Invoice 2', 'url': expense.upload_invoice_transport_2})
        if expense.payment_summary_invoice:
            attachments.append({'label': 'Payment Summary Invoice', 'url': expense.payment_summary_invoice})
        # Operation invoices
        if expense.upload_invoice_operation_1:
            attachments.append({'label': 'Operation Invoice 1', 'url': expense.upload_invoice_operation_1})
        if expense.upload_invoice_operation_2:
            attachments.append({'label': 'Operation Invoice 2', 'url': expense.upload_invoice_operation_2})
        # Stationary invoices
        if expense.upload_invoice_stationary_1:
            attachments.append({'label': 'Stationary Invoice 1', 'url': expense.upload_invoice_stationary_1})
        if expense.upload_invoice_stationary_2:
            attachments.append({'label': 'Stationary Invoice 2', 'url': expense.upload_invoice_stationary_2})
        # Other invoices
        if expense.upload_invoice_other_1:
            attachments.append({'label': 'Other Invoice 1', 'url': expense.upload_invoice_other_1})
        if expense.upload_invoice_other_2:
            attachments.append({'label': 'Other Invoice 2', 'url': expense.upload_invoice_other_2})

        # Helper function to format decimal
        def format_decimal(value):
            return float(value) if value else None

        return JsonResponse({
            # Basic Info
            'id': expense.id,
            'unique_expense_number': expense.unique_expense_number,
            'timestamp_display': timestamp_display,
            'submitted_by': expense.submitted_by,
            'email_address': expense.email_address,
            'client_name': expense.client_name,
            'client': expense.client,
            'service_month': expense.service_month,
            'nature_of_expense': expense.nature_of_expense,
            'amount': float(expense.amount) if expense.amount else 0,
            'amount_display': amount_display,
            'payment_method': expense.payment_method,
            'expenses_borne_by': expense.expenses_borne_by,
            'remark': expense.remark,
            'approval_status': expense.approval_status,
            'entered_in_tally': expense.entered_in_tally,

            # Transport Fields
            'transport': expense.transport,
            'transport_type': expense.transport_type,
            'transporter_name': expense.transporter_name,
            'from_address': expense.from_address,
            'to_address': expense.to_address,
            'vehicle_no': expense.vehicle_no,
            'invoice_no': expense.invoice_no,
            'charges_at_gw': format_decimal(expense.charges_at_gw),
            'charges_at_client': format_decimal(expense.charges_at_client),
            'unloading_box_expense': format_decimal(expense.unloading_box_expense),
            'box_count': expense.box_count,
            'warai_charges': format_decimal(expense.warai_charges),
            'labour_charges': format_decimal(expense.labour_charges),
            'pod_hard_copy': expense.pod_hard_copy,
            'expense_paid_by_transport': expense.expense_paid_by_transport,
            'mention_other_transport': expense.mention_other_transport,
            'payment_summary_invoice': expense.payment_summary_invoice,

            # Operation Fields
            'operation': expense.operation,
            'operation_expense_type': expense.operation_expense_type,
            'operation_expense_amount': format_decimal(expense.operation_expense_amount),
            'expense_paid_by_operation': expense.expense_paid_by_operation,
            'mention_other_operation': expense.mention_other_operation,

            # Stationary Fields
            'stationary': expense.stationary,
            'stationary_expense_type': expense.stationary_expense_type,
            'stationary_expense_amount': format_decimal(expense.stationary_expense_amount),
            'expense_paid_by_stationary': expense.expense_paid_by_stationary,
            'mention_other_stationary': expense.mention_other_stationary,

            # Other Expense Fields
            'other': expense.other,
            'other_expense_type': expense.other_expense_type,
            'other_expense_amount': format_decimal(expense.other_expense_amount),
            'expense_paid_by_other': expense.expense_paid_by_other,
            'mention_other_remarks': expense.mention_other_remarks,

            # Attachments
            'attachments': attachments,

            # Timestamps
            'synced_at': expense.synced_at.isoformat(),
            'updated_at': expense.updated_at.isoformat(),

            # Full raw data for debugging/complete information
            'raw_data': expense.raw_data,
        })

    except ExpenseRecord.DoesNotExist:
        return JsonResponse({'error': 'Expense not found'}, status=404)
    except Exception as e:
        logger.error(f"Failed to fetch expense details: {e}")
        return JsonResponse({'error': str(e)}, status=500)


def _trigger_sync_task(token_id, sync_type='incremental', triggered_by_user=None):
    """
    Trigger Cloud Tasks sync job.

    Args:
        token_id: GoogleSheetsToken primary key
        sync_type: 'incremental' or 'full'
        triggered_by_user: Username who triggered the sync (optional)
    """
    # If Cloud Tasks not available, run synchronously
    if not tasks_v2:
        logger.warning("Cloud Tasks not available, running sync synchronously")
        engine = ExpenseLogSyncEngine(token_id, sync_type, triggered_by_user=triggered_by_user)
        engine.sync()
        return

    try:
        client = tasks_v2.CloudTasksClient()
        project = os.getenv('GOOGLE_CLOUD_PROJECT')
        location = os.getenv('CLOUD_TASKS_LOCATION', 'asia-south1')
        queue = os.getenv('CLOUD_TASKS_QUEUE', 'default')

        parent = client.queue_path(project, location, queue)

        # Task payload
        payload = {
            'token_id': token_id,
            'sync_type': sync_type,
            'triggered_by_user': triggered_by_user,
        }

        # Create task
        task = {
            'http_request': {
                'http_method': tasks_v2.HttpMethod.POST,
                'url': f"{os.getenv('APP_URL')}/expense-log/worker/sync/",
                'headers': {
                    'Content-Type': 'application/json',
                },
                'body': json.dumps(payload).encode(),
            }
        }

        # Schedule task
        response = client.create_task(request={'parent': parent, 'task': task})
        logger.info(f"Created sync task: {response.name}")

    except Exception as e:
        logger.error(f"Failed to create sync task: {e}", exc_info=True)
        # Fallback: run sync synchronously
        logger.warning("Falling back to synchronous sync")
        engine = ExpenseLogSyncEngine(token_id, sync_type, triggered_by_user=triggered_by_user)
        engine.sync()


@login_required
def user_mappings(request):
    """
    User Name Mappings management page.
    Admin-only: Map ERP users to their Google Sheets "Submitted By" names.
    """
    # Only admins can manage mappings
    if request.user.role not in ['admin', 'director']:
        messages.error(request, "Only admins can manage user mappings")
        return redirect('expense_log:dashboard')

    from django.contrib.auth import get_user_model
    User = get_user_model()

    # Get only active users with their mappings
    users_data = []
    all_users = User.objects.filter(is_active=True).order_by('username')

    for user in all_users:
        try:
            mapping = UserNameMapping.objects.get(erp_user=user)
        except UserNameMapping.DoesNotExist:
            mapping = None

        users_data.append({
            'user': user,
            'mapping': mapping
        })

    # Calculate stats
    total_users = all_users.count()
    mapped_users = UserNameMapping.objects.count()
    unmapped_users = total_users - mapped_users

    # Get distinct sheet names from all expense records
    distinct_sheet_names = ExpenseRecord.objects.exclude(
        submitted_by__isnull=True
    ).exclude(
        submitted_by=''
    ).values_list('submitted_by', flat=True).distinct().order_by('submitted_by')

    # Add "ALL NAMES" option at the beginning
    sheet_names_with_all = ['ALL_NAMES'] + list(distinct_sheet_names)

    context = {
        'users': users_data,
        'total_users': total_users,
        'mapped_users': mapped_users,
        'unmapped_users': unmapped_users,
        'distinct_sheet_names': sheet_names_with_all,
    }

    return render(request, 'expense_log/user_mappings.html', context)


@login_required
@require_http_methods(['POST'])
def update_mapping(request, user_id):
    """
    Create or update a single user's mapping.
    Supports multiple sheet names.
    """
    if request.user.role not in ['admin', 'director']:
        messages.error(request, "Permission denied")
        return redirect('expense_log:dashboard')

    from django.contrib.auth import get_user_model
    User = get_user_model()

    try:
        user = User.objects.get(pk=user_id)
        # Get list of sheet names from multi-select
        sheet_names = request.POST.getlist('sheet_name')
        # Filter out empty values
        sheet_names = [name.strip() for name in sheet_names if name.strip()]

        if not sheet_names:
            # Delete mapping if no sheet names selected
            UserNameMapping.objects.filter(erp_user=user).delete()
            messages.success(request, f"Removed mapping for {user.get_full_name() or user.username}")
        else:
            # Create or update mapping with list of sheet names
            mapping, created = UserNameMapping.objects.update_or_create(
                erp_user=user,
                defaults={
                    'sheet_name': sheet_names,
                    'created_by': request.user
                }
            )
            action = "Created" if created else "Updated"
            names_display = ', '.join(sheet_names)
            messages.success(request, f"{action} mapping for {user.get_full_name() or user.username} → {names_display}")

    except User.DoesNotExist:
        messages.error(request, "User not found")

    return redirect('expense_log:user_mappings')


@login_required
@require_http_methods(['POST'])
def delete_mapping(request, user_id):
    """
    Delete a user's mapping.
    """
    if request.user.role not in ['admin', 'director']:
        messages.error(request, "Permission denied")
        return redirect('expense_log:dashboard')

    from django.contrib.auth import get_user_model
    User = get_user_model()

    try:
        user = User.objects.get(pk=user_id)
        deleted_count, _ = UserNameMapping.objects.filter(erp_user=user).delete()

        if deleted_count > 0:
            messages.success(request, f"Removed mapping for {user.get_full_name() or user.username}")
        else:
            messages.warning(request, f"No mapping found for {user.get_full_name() or user.username}")

    except User.DoesNotExist:
        messages.error(request, "User not found")

    return redirect('expense_log:user_mappings')


@login_required
def transport_expenses_projectwise(request):
    """
    Transport expenses grouped by (client_name, transporter_name).
    Only Approved expenses. Shows Charges@GW, Charges@Client, Margin%.
    """
    # Base queryset: transport expenses, Approved only
    transport_expenses = ExpenseRecord.get_expenses_for_user(request.user).filter(
        Q(nature_of_expense__icontains='transport') |
        Q(raw_data__Transport__isnull=False)
    ).exclude(
        client_name__isnull=True
    ).exclude(
        client_name=''
    ).filter(
        approval_status='Approved'
    )

    # Filters — default to current calendar month
    current_month = datetime.now().strftime('%B %Y')

    month_filter = request.GET.get('month', current_month)
    search_query = request.GET.get('search', '').strip()

    if month_filter:
        transport_expenses = transport_expenses.filter(service_month=month_filter)

    if search_query:
        transport_expenses = transport_expenses.filter(
            Q(client_name__icontains=search_query) |
            Q(transporter_name__icontains=search_query)
        )

    # Group by (client_name, transporter_name) using DB-level aggregation
    # Only include rows that have actual charges (exclude zero/null charge records)
    grouped = transport_expenses.filter(
        Q(charges_at_gw__isnull=False) | Q(charges_at_client__isnull=False)
    ).values(
        'client_name', 'transporter_name'
    ).annotate(
        charges_at_gw=Coalesce(Sum('charges_at_gw'), Decimal('0')),
        charges_at_client=Coalesce(Sum('charges_at_client'), Decimal('0')),
    ).filter(
        Q(charges_at_gw__gt=0) | Q(charges_at_client__gt=0)
    ).order_by('client_name', 'transporter_name')

    rows = []
    for row in grouped:
        gw = row['charges_at_gw'] or Decimal('0')
        client = row['charges_at_client'] or Decimal('0')
        transporter = row['transporter_name'] or '-'
        margin = ((client - gw) / gw * 100) if gw > 0 else Decimal('0')
        rows.append({
            'client_name': row['client_name'],
            'transporter_name': transporter,
            'charges_at_gw': gw,
            'charges_at_client': client,
            'margin': margin,
        })

    # Sort by actual client name from ProjectCode table, fall back to raw client_name
    from projects.models import ProjectCode
    project_client_names = dict(
        ProjectCode.objects.values_list('project_code', 'client_name')
    )
    rows.sort(key=lambda r: (
        project_client_names.get(r['client_name'], r['client_name'] or ''),
        r['transporter_name'] or ''
    ))

    # Totals
    total_gw = sum((r['charges_at_gw'] for r in rows), Decimal('0'))
    total_client = sum((r['charges_at_client'] for r in rows), Decimal('0'))
    total_margin = ((total_client - total_gw) / total_gw * 100) if total_gw > 0 else Decimal('0')
    total_rows = len(rows)

    # Distinct months for dropdown (all transport expenses, not filtered)
    distinct_months = ExpenseRecord.get_expenses_for_user(request.user).filter(
        Q(nature_of_expense__icontains='transport') |
        Q(raw_data__Transport__isnull=False)
    ).exclude(
        service_month__isnull=True
    ).exclude(
        service_month=''
    ).values_list('service_month', flat=True).distinct().order_by('-service_month')

    context = {
        'rows': rows,
        'total_gw': total_gw,
        'total_client': total_client,
        'total_margin': total_margin,
        'total_rows': total_rows,
        'month_filter': month_filter,
        'search_query': search_query,
        'distinct_months': distinct_months,
        'current_month': current_month,
    }

    return render(request, 'expense_log/transport_projectwise.html', context)


@login_required
def transport_projectwise_excel(request):
    """
    Download Transport Expenses - Project-wise as a formatted Excel file.
    Sheet 1: Summary with hyperlinks to per-project sheets.
    Sheet per client: individual expense records rendered as visual card blocks.
    """
    # ── Style helpers ──────────────────────────────────────────────────────────
    TEAL_HEX       = "0D9488"
    LIGHT_TEAL_HEX = "CCFBF1"
    GRAY_HEX       = "F3F4F6"
    WHITE_HEX      = "FFFFFF"
    GREEN_HEX      = "059669"
    RED_HEX        = "DC2626"
    DARK_HEX       = "1F2937"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, color=DARK_HEX, size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic)

    thin_side = Side(style="thin", color="D1D5DB")
    thick_side = Side(style="medium", color="9CA3AF")

    def thin_border():
        return Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def thick_border():
        return Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)

    def center_align(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def left_align(wrap=False):
        return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

    # ── Queryset (same logic as transport_expenses_projectwise) ──────────────
    current_month = datetime.now().strftime('%B %Y')
    month_filter  = request.GET.get('month', current_month)
    search_query  = request.GET.get('search', '').strip()

    base_qs = ExpenseRecord.get_expenses_for_user(request.user).filter(
        Q(nature_of_expense__icontains='transport') |
        Q(raw_data__Transport__isnull=False)
    ).exclude(client_name__isnull=True).exclude(client_name='').filter(approval_status='Approved')

    if month_filter:
        base_qs = base_qs.filter(service_month=month_filter)
    if search_query:
        base_qs = base_qs.filter(
            Q(client_name__icontains=search_query) |
            Q(transporter_name__icontains=search_query)
        )

    grouped = base_qs.filter(
        Q(charges_at_gw__isnull=False) | Q(charges_at_client__isnull=False)
    ).values('client_name', 'transporter_name').annotate(
        charges_at_gw=Coalesce(Sum('charges_at_gw'), Decimal('0')),
        charges_at_client=Coalesce(Sum('charges_at_client'), Decimal('0')),
    ).filter(
        Q(charges_at_gw__gt=0) | Q(charges_at_client__gt=0)
    ).order_by('client_name', 'transporter_name')

    rows = []
    for row in grouped:
        gw      = row['charges_at_gw'] or Decimal('0')
        client  = row['charges_at_client'] or Decimal('0')
        margin  = ((client - gw) / gw * 100) if gw > 0 else Decimal('0')
        rows.append({
            'client_name':     row['client_name'],
            'transporter_name': row['transporter_name'] or '-',
            'charges_at_gw':   gw,
            'charges_at_client': client,
            'margin':          margin,
        })

    total_gw     = sum((r['charges_at_gw'] for r in rows), Decimal('0'))
    total_client = sum((r['charges_at_client'] for r in rows), Decimal('0'))
    total_margin = ((total_client - total_gw) / total_gw * 100) if total_gw > 0 else Decimal('0')

    # ── Sheet name sanitization ───────────────────────────────────────────────
    def safe_sheet_name(name, used):
        clean = re.sub(r'[\[\]:*?/\\]', '', name or 'Unknown')[:31]
        original = clean
        idx = 1
        while clean in used:
            suffix = f"_{idx}"
            clean  = original[:31 - len(suffix)] + suffix
            idx   += 1
        used.add(clean)
        return clean

    # Build sheet name map (client_name → sheet_name) for hyperlinks
    used_names   = {"Summary"}
    client_sheets = {}
    for row in rows:
        cn = row['client_name']
        if cn not in client_sheets:
            client_sheets[cn] = safe_sheet_name(cn, used_names)

    # ── Workbook ──────────────────────────────────────────────────────────────
    # ── Auto column width helper ──────────────────────────────────────────────
    def auto_fit_columns(sheet, min_width=10, max_width=60, padding=2):
        col_widths = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                col = get_column_letter(cell.column)
                try:
                    length = len(str(cell.value))
                except Exception:
                    length = 10
                col_widths[col] = min(max_width, max(col_widths.get(col, min_width), length + padding))
        for col, width in col_widths.items():
            sheet.column_dimensions[col].width = width

    wb = openpyxl.Workbook()

    # ══ SHEET 1: Summary ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"

    # Title row
    ws.merge_cells('A1:E1')
    c = ws['A1']
    c.value         = "Transport Expenses – Project-wise"
    c.font          = font(bold=True, color=WHITE_HEX, size=14)
    c.fill          = fill(TEAL_HEX)
    c.alignment     = center_align()
    ws.row_dimensions[1].height = 28

    # Filter info row
    filter_info = f"Month: {month_filter or 'All'}  |  Search: {search_query or 'All'}"
    ws.merge_cells('A2:E2')
    c = ws['A2']
    c.value     = filter_info
    c.font      = font(italic=True, color="6B7280", size=9)
    c.alignment = center_align()
    ws.row_dimensions[2].height = 16

    # KPI boxes (row 4 = labels, row 5 = values)
    ws.row_dimensions[3].height = 8
    kpi_labels  = ["Total Charges@GW", "Total Charges@Client", "Overall Margin"]
    kpi_values  = [float(total_gw), float(total_client), float(total_margin)]
    kpi_formats = ['₹#,##0.00', '₹#,##0.00', '0.00"%"']
    kpi_cols    = ['B', 'C', 'D']
    for col, label, val, fmt in zip(kpi_cols, kpi_labels, kpi_values, kpi_formats):
        lc = ws[f'{col}4']
        lc.value     = label
        lc.font      = font(bold=True, color="374151", size=9)
        lc.fill      = fill(LIGHT_TEAL_HEX)
        lc.alignment = center_align()
        lc.border    = thin_border()

        vc = ws[f'{col}5']
        vc.value         = val
        vc.number_format = fmt
        vc.font          = font(bold=True, color=TEAL_HEX, size=12)
        vc.fill          = fill(WHITE_HEX)
        vc.alignment     = center_align()
        vc.border        = thin_border()

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 8

    # Summary table header (row 7)
    headers = ["Client", "Transporter", "Charges@GW", "Charges@Client", "Margin%"]
    for col_idx, hdr in enumerate(headers, 1):
        c = ws.cell(row=7, column=col_idx, value=hdr)
        c.font      = font(bold=True, color=WHITE_HEX, size=10)
        c.fill      = fill(TEAL_HEX)
        c.alignment = center_align()
        c.border    = thin_border()
    ws.row_dimensions[7].height = 20
    ws.freeze_panes = 'A8'

    # Total row (row 8)
    total_row_data = ["TOTAL", "", float(total_gw), float(total_client), float(total_margin)]
    total_fmts     = [None, None, '₹#,##0.00', '₹#,##0.00', '0.00"%"']
    for col_idx, (val, fmt) in enumerate(zip(total_row_data, total_fmts), 1):
        c = ws.cell(row=8, column=col_idx, value=val)
        c.font      = font(bold=True, color=DARK_HEX, size=10)
        c.fill      = fill(GRAY_HEX)
        c.alignment = center_align() if col_idx > 2 else left_align()
        c.border    = thin_border()
        if fmt:
            c.number_format = fmt

    # Data rows
    for row_num, row in enumerate(rows, 9):
        ws.row_dimensions[row_num].height = 18
        cn       = row['client_name']
        sheet_nm = client_sheets.get(cn, cn)
        margin   = float(row['margin'])

        # Col A: Client with hyperlink to project sheet
        c = ws.cell(row=row_num, column=1, value=cn)
        c.hyperlink = f"#'{sheet_nm}'!A1"
        c.font      = Font(bold=False, color="0D9488", size=10, underline="single")
        c.alignment = left_align()
        c.border    = thin_border()

        # Col B: Transporter
        c = ws.cell(row=row_num, column=2, value=row['transporter_name'])
        c.font      = font()
        c.alignment = left_align()
        c.border    = thin_border()

        # Col C: Charges@GW
        c = ws.cell(row=row_num, column=3, value=float(row['charges_at_gw']))
        c.number_format = '₹#,##0.00'
        c.font          = font()
        c.alignment     = center_align()
        c.border        = thin_border()

        # Col D: Charges@Client
        c = ws.cell(row=row_num, column=4, value=float(row['charges_at_client']))
        c.number_format = '₹#,##0.00'
        c.font          = font()
        c.alignment     = center_align()
        c.border        = thin_border()

        # Col E: Margin%
        margin_color = GREEN_HEX if margin >= 0 else RED_HEX
        c = ws.cell(row=row_num, column=5, value=margin / 100)
        c.number_format = '0.00%'
        c.font          = font(bold=True, color=margin_color)
        c.alignment     = center_align()
        c.border        = thin_border()

    auto_fit_columns(ws)

    # ══ Per-client sheets ═════════════════════════════════════════════════════
    for cn, sheet_nm in client_sheets.items():
        ws2 = wb.create_sheet(title=sheet_nm)

        # Back link row 1
        back = ws2['A1']
        back.value     = "← Back to Summary"
        back.hyperlink = "#'Summary'!A1"
        back.font      = Font(color="0D9488", size=10, underline="single", bold=True)
        back.alignment = left_align()
        ws2.row_dimensions[1].height = 18

        # Sheet title row 2
        ws2.merge_cells('A2:B2')
        c = ws2['A2']
        c.value     = f"Transport Records — {cn}"
        c.font      = font(bold=True, color=WHITE_HEX, size=12)
        c.fill      = fill(TEAL_HEX)
        c.alignment = center_align()
        ws2.row_dimensions[2].height = 24

        # Fetch individual records for this client
        records = base_qs.filter(client_name=cn).order_by('-timestamp')

        current_row = 4  # start writing cards from row 4

        # Track charge totals for subtotal
        sub_gw     = Decimal('0')
        sub_client = Decimal('0')

        for exp in records:
            gw_val     = exp.charges_at_gw or Decimal('0')
            client_val = exp.charges_at_client or Decimal('0')
            sub_gw    += gw_val
            sub_client += client_val
            margin_val = float(((client_val - gw_val) / gw_val * 100)) if gw_val > 0 else 0.0

            # Status color
            status_color = GREEN_HEX if exp.approval_status == 'Approved' else (
                RED_HEX if exp.approval_status == 'Rejected' else "D97706"
            )

            # ── Card header row ──────────────────────────────────────────────
            ws2.merge_cells(f'A{current_row}:B{current_row}')
            c = ws2[f'A{current_row}']
            uen_str    = exp.unique_expense_number or '-'
            month_str  = exp.service_month or '-'
            status_str = exp.approval_status or 'Pending'
            c.value     = f"  {uen_str}    |    {month_str}    |    {status_str}"
            c.font      = Font(bold=True, color=WHITE_HEX, size=10)
            c.fill      = fill(TEAL_HEX)
            c.alignment = center_align()
            c.border    = thin_border()
            ws2.row_dimensions[current_row].height = 20
            current_row += 1

            def card_section(label):
                nonlocal current_row
                ws2.merge_cells(f'A{current_row}:B{current_row}')
                c = ws2[f'A{current_row}']
                c.value     = label
                c.font      = font(bold=True, color="0F766E", size=9)
                c.fill      = fill(LIGHT_TEAL_HEX)
                c.alignment = left_align()
                c.border    = thin_border()
                ws2.row_dimensions[current_row].height = 16
                current_row += 1

            def card_field(label, value, fmt=None, value_color=None):
                nonlocal current_row
                lc = ws2[f'A{current_row}']
                lc.value     = label
                lc.font      = font(bold=False, color="4B5563", size=9)
                lc.alignment = left_align()
                lc.border    = thin_border()

                vc = ws2[f'B{current_row}']
                vc.value     = value if value not in (None, '', '-') else '-'
                vc.font      = font(bold=False, color=value_color or DARK_HEX, size=10)
                vc.alignment = left_align(wrap=True)
                vc.border    = thin_border()
                if fmt:
                    vc.number_format = fmt
                ws2.row_dimensions[current_row].height = 16
                current_row += 1

            # ── Basic Information ────────────────────────────────────────────
            card_section("📋  BASIC INFORMATION")
            card_field("Submitted By",      exp.submitted_by)
            card_field("Service Month",     exp.service_month)
            card_field("Payment Method",    exp.payment_method)
            card_field("Expenses Borne By", exp.expenses_borne_by)

            # ── Transport Info ───────────────────────────────────────────────
            card_section("🚚  TRANSPORT INFO")
            card_field("Transport Type", exp.transport_type)
            card_field("Transporter",    exp.transporter_name)
            card_field("Vehicle No",     exp.vehicle_no)
            card_field("Invoice No",     exp.invoice_no)

            # ── Route ────────────────────────────────────────────────────────
            card_section("🗺  ROUTE")
            card_field("From", exp.from_address)
            card_field("To",   exp.to_address)

            # ── Charges ──────────────────────────────────────────────────────
            card_section("💰  CHARGES")
            card_field("Charges @ GW",     float(gw_val) if gw_val else '-',     fmt='₹#,##0.00')
            card_field("Charges @ Client", float(client_val) if client_val else '-', fmt='₹#,##0.00')
            card_field("Margin",           margin_val / 100 if gw_val > 0 else '-',
                       fmt='0.00%',
                       value_color=GREEN_HEX if margin_val >= 0 else RED_HEX)

            # ── Additional Charges ───────────────────────────────────────────
            card_section("💵  ADDITIONAL CHARGES")
            card_field("Unloading Box Expense", float(exp.unloading_box_expense) if exp.unloading_box_expense else '-', fmt='₹#,##0.00')
            card_field("Warai Charges",         float(exp.warai_charges) if exp.warai_charges else '-',         fmt='₹#,##0.00')
            card_field("Labour Charges",        float(exp.labour_charges) if exp.labour_charges else '-',        fmt='₹#,##0.00')
            card_field("Box Count",             exp.box_count)
            card_field("POD Hard Copy",         exp.pod_hard_copy)
            card_field("Paid By (Transport)",   exp.expense_paid_by_transport)

            # ── Remarks ──────────────────────────────────────────────────────
            card_section("📝  REMARKS")
            card_field("Other Transport", exp.mention_other_transport)
            card_field("General Remark",  exp.remark or None)

            # ── Attachments ──────────────────────────────────────────────────
            attachments = []
            if exp.transport_bill:
                attachments.append(("Transport Bill", exp.transport_bill))
            if exp.upload_invoice_transport_2:
                attachments.append(("Transport Invoice 2", exp.upload_invoice_transport_2))
            if exp.payment_summary_invoice:
                attachments.append(("Payment Summary Invoice", exp.payment_summary_invoice))

            if attachments:
                card_section("📎  ATTACHMENTS")
                for att_label, att_url in attachments:
                    lc = ws2[f'A{current_row}']
                    lc.value     = att_label
                    lc.font      = font(bold=False, color="4B5563", size=9)
                    lc.alignment = left_align()
                    lc.border    = thin_border()

                    vc = ws2[f'B{current_row}']
                    vc.value     = att_url
                    vc.hyperlink = att_url
                    vc.font      = Font(color="2563EB", size=9, underline="single")
                    vc.alignment = left_align(wrap=True)
                    vc.border    = thin_border()
                    ws2.row_dimensions[current_row].height = 16
                    current_row += 1

            # blank row between cards
            ws2.row_dimensions[current_row].height = 10
            current_row += 1

        auto_fit_columns(ws2, min_width=18, max_width=55)

        # Subtotal row at bottom
        if records.exists():
            sub_margin = float(((sub_client - sub_gw) / sub_gw * 100)) if sub_gw > 0 else 0.0
            current_row += 1
            ws2.merge_cells(f'A{current_row}:B{current_row}')
            c = ws2[f'A{current_row}']
            c.value = (
                f"SUBTOTAL  |  Charges@GW: ₹{float(sub_gw):,.2f}  "
                f"|  Charges@Client: ₹{float(sub_client):,.2f}  "
                f"|  Margin: {sub_margin:.2f}%"
            )
            c.font      = font(bold=True, color=DARK_HEX, size=10)
            c.fill      = fill(GRAY_HEX)
            c.alignment = center_align()
            c.border    = thin_border()
            ws2.row_dimensions[current_row].height = 22
        else:
            ws2.merge_cells(f'A{current_row}:B{current_row}')
            c = ws2[f'A{current_row}']
            c.value     = "No expense records found for this project."
            c.font      = font(italic=True, color="9CA3AF", size=10)
            c.alignment = center_align()

    # ── Stream response ───────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_month = re.sub(r'[^A-Za-z0-9_-]', '_', month_filter or 'All')
    filename   = f"Transport_Expenses_{safe_month}.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def transport_project_detail(request, project_code):
    """
    Detail view for a specific project's transport expenses.
    Shows all UENs and expense details for the project.
    """
    from django.db.models import Q
    from datetime import datetime

    # Get user's expenses
    all_expenses = ExpenseRecord.get_expenses_for_user(request.user).select_related('token')

    # Filter for this project's transport expenses
    expenses = all_expenses.filter(
        Q(nature_of_expense__icontains='transport') |
        Q(raw_data__Transport__isnull=False),
        client_name=project_code
    )

    # Filters - preserve from main page
    status_filter = request.GET.get('status', '')
    current_month = datetime.now().strftime('%B %Y')  # e.g., "February 2026"
    month_filter = request.GET.get('month', current_month)
    transporter_filter = request.GET.get('transporter', '')

    if status_filter:
        expenses = expenses.filter(approval_status=status_filter)

    if month_filter:
        expenses = expenses.filter(service_month=month_filter)

    if transporter_filter:
        expenses = expenses.filter(transporter_name=transporter_filter)

    expenses = expenses.order_by('-timestamp')

    # Get project info
    from projects.models import ProjectCode
    project = None
    try:
        project = ProjectCode.objects.get(code=project_code)
    except ProjectCode.DoesNotExist:
        pass

    # Get distinct months for this project
    distinct_months = all_expenses.filter(
        Q(nature_of_expense__icontains='transport') |
        Q(raw_data__Transport__isnull=False),
        client_name=project_code
    ).exclude(
        service_month__isnull=True
    ).exclude(
        service_month=''
    ).values_list('service_month', flat=True).distinct().order_by('-service_month')

    context = {
        'project_code': project_code,
        'project': project,
        'expenses': expenses,
        'status_filter': status_filter,
        'month_filter': month_filter,
        'transporter_filter': transporter_filter,
        'distinct_months': distinct_months,
        'current_month': current_month,
    }

    return render(request, 'expense_log/transport_project_detail.html', context)
