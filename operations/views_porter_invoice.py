"""
Porter Invoice Editor Views
Batch processing + single invoice editing for Porter transport invoices.
"""

import io
import json
import logging
import math
import re
import tempfile
import zipfile
from decimal import Decimal
from pathlib import Path

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.http import FileResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models_porter_invoice import PorterInvoiceSession, PorterInvoiceFile

logger = logging.getLogger(__name__)

# Roles allowed for Porter Invoice Editor
PORTER_ALLOWED_ROLES = [
    'operation_coordinator', 'operation_manager', 'operation_controller',
    'admin', 'super_user',
]

_CRN_PREFIX_RE = re.compile(r"^CRN", re.IGNORECASE)
CRN_HEADER_CANDIDATES = (
    'crn',
    'crn no',
    'crn number',
    'order id',
    'order no',
    'order number',
    'invoice id',
    'invoice no',
    'invoice number',
)
TOTAL_HEADER_CANDIDATES = (
    'new total',
    'total amount',
    'total',
    'amount',
    'price',
    'fare',
)
EXCEL_VALIDATION_SAMPLE_SIZE = 20
EXCEL_VALIDATION_THRESHOLD = 0.90


def _check_access(request):
    """Return redirect response if user lacks access, else None."""
    if request.user.role not in PORTER_ALLOWED_ROLES:
        messages.error(request, "You don't have access to the Porter Invoice Editor.")
        return redirect('accounts:dashboard')
    return None


def _normalise_crn(raw: str) -> str:
    """Strip CRN prefix and Excel float-suffix so keys always match."""
    s = str(raw).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return _CRN_PREFIX_RE.sub("", s)


def _normalise_excel_header(value) -> str:
    """Normalise Excel header text for safer column matching."""
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _find_excel_column_index(header, candidates):
    """Return the first header index matching any candidate phrase, else None."""
    for i, h in enumerate(header):
        if any(candidate in h for candidate in candidates):
            return i
    return None


def _looks_like_crn(value) -> bool:
    """Return True when the cell value matches the expected CRN format."""
    text = str(value or '').strip().upper()
    if text.endswith('.0'):
        text = text[:-2]
    return text.startswith('CRN') and len(text) > 3


def _looks_like_amount(value) -> bool:
    """Return True when the cell value can be parsed as a decimal amount."""
    text = str(value or '').strip().replace(',', '')
    if not text:
        return False
    try:
        Decimal(text)
        return True
    except Exception:
        return False


def _parse_excel_mapping(excel_file, crn_col_index=None, total_col_index=None) -> dict:
    """Parse Excel file to get CRN → target_total mapping using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [_normalise_excel_header(h) for h in rows[0]]

    crn_col = crn_col_index if crn_col_index is not None else _find_excel_column_index(
        header, CRN_HEADER_CANDIDATES,
    )
    total_col = total_col_index if total_col_index is not None else _find_excel_column_index(
        header, TOTAL_HEADER_CANDIDATES,
    )

    if (
        crn_col is None or total_col is None
        or crn_col < 0 or total_col < 0
        or crn_col >= len(header) or total_col >= len(header)
    ):
        return {}

    mapping = {}
    for row in rows[1:]:
        if len(row) <= max(crn_col, total_col):
            continue
        crn_val = row[crn_col]
        total_val = row[total_col]
        if crn_val and total_val:
            key = _normalise_crn(str(crn_val))
            try:
                mapping[key] = Decimal(str(total_val).replace(',', ''))
            except Exception:
                continue

    wb.close()
    return mapping


def _validate_excel_crn_column(excel_file, requested_col_index=None):
    """Validate that the chosen CRN column mostly contains CRN-like values."""
    import openpyxl

    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return False, None, [], 0, 0

    header = [_normalise_excel_header(h) for h in rows[0]]
    crn_col = requested_col_index if requested_col_index is not None else _find_excel_column_index(
        header, CRN_HEADER_CANDIDATES,
    )

    if crn_col is None or crn_col < 0 or crn_col >= len(header):
        wb.close()
        return False, crn_col, [], 0, 0

    samples = []
    for row in rows[1:]:
        if len(row) <= crn_col:
            continue
        value = row[crn_col]
        if value in (None, ''):
            continue
        samples.append(str(value).strip())
        if len(samples) >= EXCEL_VALIDATION_SAMPLE_SIZE:
            break

    wb.close()

    if not samples:
        return False, crn_col, [], 0, 0

    valid_count = sum(1 for value in samples if _looks_like_crn(value))
    required_count = max(1, math.ceil(len(samples) * EXCEL_VALIDATION_THRESHOLD))
    return valid_count >= required_count, crn_col, samples, valid_count, len(samples)


def _validate_excel_total_column(excel_file, requested_col_index=None):
    """Validate that the chosen total column mostly contains numeric values."""
    import openpyxl

    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return False, None, [], 0, 0

    header = [_normalise_excel_header(h) for h in rows[0]]
    total_col = requested_col_index if requested_col_index is not None else _find_excel_column_index(
        header, TOTAL_HEADER_CANDIDATES,
    )

    if total_col is None or total_col < 0 or total_col >= len(header):
        wb.close()
        return False, total_col, [], 0, 0

    samples = []
    for row in rows[1:]:
        if len(row) <= total_col:
            continue
        value = row[total_col]
        if value in (None, ''):
            continue
        samples.append(str(value).strip())
        if len(samples) >= EXCEL_VALIDATION_SAMPLE_SIZE:
            break

    wb.close()

    if not samples:
        return False, total_col, [], 0, 0

    valid_count = sum(1 for value in samples if _looks_like_amount(value))
    required_count = max(1, math.ceil(len(samples) * EXCEL_VALIDATION_THRESHOLD))
    return valid_count >= required_count, total_col, samples, valid_count, len(samples)


# ──────────────────────────────────────────────────────────────
# 1. Dashboard — session list
# ──────────────────────────────────────────────────────────────

@login_required
def porter_invoice_dashboard(request):
    denied = _check_access(request)
    if denied:
        return denied

    sessions = PorterInvoiceSession.objects.select_related('created_by')

    # Filters
    session_type = request.GET.get('type', '')
    if session_type in ('batch', 'single'):
        sessions = sessions.filter(session_type=session_type)

    status_filter = request.GET.get('status', '')
    if status_filter:
        sessions = sessions.filter(status=status_filter)

    paginator = Paginator(sessions, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Summary stats
    total_sessions = PorterInvoiceSession.objects.count()
    total_files_processed = PorterInvoiceFile.objects.filter(status='success').count()

    context = {
        'page_obj': page_obj,
        'session_type': session_type,
        'status_filter': status_filter,
        'total_sessions': total_sessions,
        'total_files_processed': total_files_processed,
    }
    return render(request, 'operations/porter_invoice_dashboard.html', context)


# ──────────────────────────────────────────────────────────────
# 2. Batch upload + process
# ──────────────────────────────────────────────────────────────

@login_required
def porter_invoice_batch(request):
    denied = _check_access(request)
    if denied:
        return denied

    if request.method == 'GET':
        return render(request, 'operations/porter_invoice_batch.html')

    def _parse_column_index(value):
        value = (value or '').strip()
        if value == '':
            return None
        try:
            return int(value)
        except ValueError:
            return None

    excel_crn_column = _parse_column_index(request.POST.get('excel_crn_column'))
    excel_total_column = _parse_column_index(request.POST.get('excel_total_column'))

    # POST — process batch
    files = request.FILES.getlist('pdf_files')
    if not files:
        messages.error(request, "Please upload at least one PDF file.")
        return render(request, 'operations/porter_invoice_batch.html')

    multiplier_str = request.POST.get('multiplier', '1.20')
    try:
        multiplier = float(multiplier_str)
    except ValueError:
        multiplier = 1.20

    # Parse optional Excel mapping
    excel_mapping = {}
    excel_file = request.FILES.get('excel_file')
    if excel_file:
        try:
            excel_file.seek(0)
            crn_is_valid, effective_crn_column, crn_samples, crn_valid_count, crn_sample_count = _validate_excel_crn_column(
                excel_file,
                requested_col_index=excel_crn_column,
            )
            if not crn_is_valid:
                sample_preview = ", ".join(crn_samples[:5]) if crn_samples else "No non-empty sample values found"
                messages.error(
                    request,
                    "Selected CRN column failed validation. "
                    f"{crn_valid_count}/{crn_sample_count} sampled values start with CRN "
                    f"(minimum {math.ceil(max(crn_sample_count, 1) * EXCEL_VALIDATION_THRESHOLD)} required). "
                    f"Preview: {sample_preview}"
                )
                return render(request, 'operations/porter_invoice_batch.html')

            excel_file.seek(0)
            total_is_valid, effective_total_column, total_samples, total_valid_count, total_sample_count = _validate_excel_total_column(
                excel_file,
                requested_col_index=excel_total_column,
            )
            if not total_is_valid:
                sample_preview = ", ".join(total_samples[:5]) if total_samples else "No non-empty sample values found"
                messages.error(
                    request,
                    "Selected New Total column failed validation. "
                    f"{total_valid_count}/{total_sample_count} sampled values look numeric "
                    f"(minimum {math.ceil(max(total_sample_count, 1) * EXCEL_VALIDATION_THRESHOLD)} required). "
                    f"Preview: {sample_preview}"
                )
                return render(request, 'operations/porter_invoice_batch.html')

            excel_file.seek(0)
            excel_mapping = _parse_excel_mapping(
                excel_file,
                crn_col_index=effective_crn_column,
                total_col_index=effective_total_column,
            )
            if not excel_mapping:
                messages.warning(
                    request,
                    "Excel mapping was uploaded but no CRN/Total mapping rows could be parsed."
                )
        except Exception as e:
            messages.warning(request, f"Failed to parse Excel file: {e}. Processing without mapping.")

    # Create session
    session = PorterInvoiceSession.objects.create(
        session_type='batch',
        status='processing',
        multiplier=Decimal(str(multiplier)),
        created_by=request.user,
        total_files=len(files),
    )

    # Save Excel mapping if provided
    if excel_file:
        excel_file.seek(0)
        session.excel_mapping_file.save(excel_file.name, excel_file, save=True)

    # Process each PDF
    from .porter_invoice.batch_runner import process_single, extract_order_number

    success_count = 0
    error_count = 0
    skipped_count = 0
    edited_files = []  # (filename, bytes) for ZIP

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        output_dir = tmp_path / 'output'
        output_dir.mkdir()

        for uploaded_file in files:
            # Normalise filename — some browsers send full paths
            safe_name = Path(uploaded_file.name).name

            # Save uploaded PDF to temp
            input_path = tmp_path / safe_name
            with open(input_path, 'wb') as f:
                for chunk in uploaded_file.chunks():
                    f.write(chunk)

            # Determine target total from Excel mapping
            order_number = extract_order_number(safe_name)
            normalised = _normalise_crn(order_number)
            target_total = excel_mapping.get(normalised)

            # Process
            result = process_single(
                pdf_path=input_path,
                output_dir=output_dir,
                multiplier=multiplier,
                target_total=target_total,
            )

            # Create file record
            file_record = PorterInvoiceFile(
                session=session,
                original_filename=safe_name,
                crn=order_number,
                status=result.status,
                error_message=result.error or '',
                old_total=result.old_total,
                new_total=result.new_total,
            )

            # Save original PDF
            uploaded_file.seek(0)
            file_record.original_pdf.save(safe_name, uploaded_file, save=False)

            # Save edited PDF if successful — use new CRN-based filename
            # output_dir contains the file saved as safe_name by process_single
            edited_path = output_dir / safe_name
            if result.status == 'success':
                if edited_path.exists():
                    new_filename = f"invoice_{order_number}.pdf"
                    with open(edited_path, 'rb') as ef:
                        file_bytes = ef.read()
                    file_record.edited_pdf.save(new_filename, ContentFile(file_bytes), save=False)
                    edited_files.append((new_filename, file_bytes))
                else:
                    # Output file missing despite success status — log for debugging
                    logger.warning(
                        "Output file not found for %s (expected: %s). "
                        "Files in output_dir: %s",
                        safe_name, edited_path,
                        [p.name for p in output_dir.iterdir()] if output_dir.exists() else "dir missing"
                    )
                    file_record.status = 'error'
                    file_record.error_message = 'Output file was not written (internal error)'

            file_record.save()

            if file_record.status == 'success':
                success_count += 1
            elif file_record.status == 'skipped':
                skipped_count += 1
            else:
                error_count += 1

    # Generate ZIP of edited files
    if edited_files:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for filename, file_bytes in edited_files:
                zf.writestr(filename, file_bytes)
        zip_buffer.seek(0)
        zip_name = f"porter_invoices_{session.id}_{timezone.now():%Y%m%d_%H%M}.zip"
        session.result_zip.save(zip_name, ContentFile(zip_buffer.read()), save=False)

    # Update session
    session.status = 'completed' if error_count == 0 else 'completed'
    session.success_count = success_count
    session.error_count = error_count
    session.skipped_count = skipped_count
    session.save()

    messages.success(
        request,
        f"Batch processed: {success_count} success, {skipped_count} skipped, {error_count} errors."
    )
    return redirect('operations:porter_invoice_detail', session_id=session.id)


# ──────────────────────────────────────────────────────────────
# 3. Single edit page
# ──────────────────────────────────────────────────────────────

@login_required
def porter_invoice_single(request):
    denied = _check_access(request)
    if denied:
        return denied
    return render(request, 'operations/porter_invoice_single.html')


# ──────────────────────────────────────────────────────────────
# 4. Session detail
# ──────────────────────────────────────────────────────────────

@login_required
def porter_invoice_detail(request, session_id):
    denied = _check_access(request)
    if denied:
        return denied

    session = get_object_or_404(
        PorterInvoiceSession.objects.select_related('created_by'),
        id=session_id
    )
    files = session.files.all()

    context = {
        'session': session,
        'files': files,
    }
    return render(request, 'operations/porter_invoice_detail.html', context)


# ──────────────────────────────────────────────────────────────
# 5. Download ZIP
# ──────────────────────────────────────────────────────────────

@login_required
def porter_invoice_download_zip(request, session_id):
    denied = _check_access(request)
    if denied:
        return denied

    session = get_object_or_404(PorterInvoiceSession, id=session_id)
    if not session.result_zip or not session.result_zip.name:
        messages.error(request, "No ZIP file available for this session.")
        return redirect('operations:porter_invoice_detail', session_id=session_id)

    try:
        file_handle = session.result_zip.open('rb')
        filename = session.result_zip.name.split('/')[-1]
        response = FileResponse(file_handle, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        messages.error(request, f"Error downloading file: {e}")
        return redirect('operations:porter_invoice_detail', session_id=session_id)


# ──────────────────────────────────────────────────────────────
# 6. Download individual file
# ──────────────────────────────────────────────────────────────

@login_required
def porter_invoice_download_file(request, file_id):
    denied = _check_access(request)
    if denied:
        return denied

    file_record = get_object_or_404(PorterInvoiceFile, id=file_id)

    # Prefer edited PDF, fall back to original
    pdf_field = file_record.edited_pdf if file_record.edited_pdf and file_record.edited_pdf.name else file_record.original_pdf

    try:
        file_handle = pdf_field.open('rb')
        # Use the edited PDF's stored filename (invoice_{CRN}.pdf) if available, else original
        if file_record.edited_pdf and file_record.edited_pdf.name:
            filename = Path(file_record.edited_pdf.name).name
        else:
            filename = file_record.original_filename
        response = FileResponse(file_handle, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        messages.error(request, f"Error downloading file: {e}")
        return redirect('operations:porter_invoice_detail', session_id=file_record.session_id)


# ──────────────────────────────────────────────────────────────
# 7. AJAX: Extract invoice (single edit — phase 1)
# ──────────────────────────────────────────────────────────────

@login_required
@require_POST
def porter_invoice_extract_api(request):
    if request.user.role not in PORTER_ALLOWED_ROLES:
        return JsonResponse({'error': 'Access denied'}, status=403)

    pdf_file = request.FILES.get('pdf_file')
    if not pdf_file:
        return JsonResponse({'error': 'No PDF file uploaded'}, status=400)

    if not pdf_file.name.lower().endswith('.pdf'):
        return JsonResponse({'error': 'File must be a PDF'}, status=400)

    from .porter_invoice.single_editor import extract_crn_from_filename

    crn = extract_crn_from_filename(pdf_file.name)

    # Create session + file record
    session = PorterInvoiceSession.objects.create(
        session_type='single',
        status='pending',
        created_by=request.user,
        total_files=1,
    )

    file_record = PorterInvoiceFile.objects.create(
        session=session,
        original_filename=pdf_file.name,
        crn=crn,
        status='pending',
    )
    file_record.original_pdf.save(pdf_file.name, pdf_file, save=True)

    # Validate PDF (download from storage to temp file for GCS compatibility)
    import fitz
    try:
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp.write(file_record.original_pdf.read())
            tmp_path = Path(tmp.name)
        doc = fitz.open(str(tmp_path))
        doc.close()
        tmp_path.unlink(missing_ok=True)
    except Exception as e:
        tmp_path.unlink(missing_ok=True)
        file_record.status = 'error'
        file_record.error_message = f'Invalid or corrupted PDF: {e}'
        file_record.save()
        session.status = 'failed'
        session.error_count = 1
        session.save()
        return JsonResponse({'error': 'Invalid or corrupted PDF file.'}, status=400)

    return JsonResponse({
        'session_id': session.id,
        'file_id': file_record.id,
        'crn': crn,
        'filename': pdf_file.name,
    })


# ──────────────────────────────────────────────────────────────
# 8. AJAX: Apply single edit (phase 2)
# ──────────────────────────────────────────────────────────────

@login_required
@require_POST
def porter_invoice_edit_api(request):
    if request.user.role not in PORTER_ALLOWED_ROLES:
        return JsonResponse({'error': 'Access denied'}, status=403)

    file_id = request.POST.get('file_id')
    if not file_id:
        return JsonResponse({'error': 'Missing file_id'}, status=400)

    file_record = get_object_or_404(PorterInvoiceFile, id=file_id)
    session = file_record.session

    # Collect edit fields
    fields = {
        'crn': request.POST.get('crn', ''),
        'date': request.POST.get('date', ''),
        'vehicle_type': request.POST.get('vehicle_type', ''),
        'vehicle': request.POST.get('vehicle', ''),
        'num_stops': request.POST.get('num_stops', ''),
        'pickup_loc': request.POST.get('pickup_loc', ''),
        'pickup_date': request.POST.get('pickup_date', ''),
        'pickup_time': request.POST.get('pickup_time', ''),
        'drop_loc': request.POST.get('drop_loc', ''),
        'drop_date': request.POST.get('drop_date', ''),
        'drop_time': request.POST.get('drop_time', ''),
    }

    from .porter_invoice.single_editor import apply_single_edit

    # Download original PDF to temp file (GCS compatible)
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_in:
        tmp_in.write(file_record.original_pdf.read())
        input_path = Path(tmp_in.name)

    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_out:
        output_path = Path(tmp_out.name)

    result = apply_single_edit(input_path, output_path, fields)
    input_path.unlink(missing_ok=True)

    if result['status'] == 'success':
        # Save edited PDF
        crn = fields.get('crn', '').strip()
        effective_crn = crn or file_record.crn or Path(file_record.original_filename).stem
        output_filename = f"invoice_{effective_crn}.pdf"

        with open(output_path, 'rb') as f:
            file_record.edited_pdf.save(output_filename, ContentFile(f.read()), save=False)

        file_record.status = 'success'
        file_record.edit_fields = {k: v for k, v in fields.items() if v}
        file_record.save()

        session.status = 'completed'
        session.success_count = 1
        session.save()

        # Clean up temp file
        output_path.unlink(missing_ok=True)

        return JsonResponse({
            'status': 'success',
            'download_url': f'/operations/porter-invoices/file/{file_record.id}/download/',
        })
    else:
        file_record.status = 'error'
        file_record.error_message = result.get('error', 'Unknown error')
        file_record.save()

        session.status = 'failed'
        session.error_count = 1
        session.save()

        output_path.unlink(missing_ok=True)

        return JsonResponse({
            'status': 'error',
            'error': result.get('error', 'Unknown error'),
        }, status=500)
